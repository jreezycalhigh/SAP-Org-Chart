import openpyxl

print("Generated SAP_Account_Priority_Outreach_Hub.html successfully!")

import openpyxl

import json

import os

import warnings

warnings.filterwarnings("ignore", category=SyntaxWarning)

wb = openpyxl.load_workbook("SAP_Account_Priority_Outreach_List_v16.xlsx")

pol_ws = wb["Priority Outreach List"]

gold_ws = wb["Gold-Star (Chart-Only)"]

# Load Gold Star manual details

try:

    with open("gold_star_manual_details.json", "r", encoding="utf-8") as f:

        gs_manual = json.load(f)

except Exception:

    gs_manual = {}

contacts = {}

# 1. Load Gold Star contacts (shown on the org charts)

rows_gold = list(gold_ws.iter_rows(values_only=True))[4:]

for r in rows_gold:

    if any(x is not None for x in r):

        name = r[0]

        title = r[1]

        reports_to = r[2]

        depth = r[3]

        direct_reports_count = r[4]

        m_info = gs_manual.get(name, {})

        contacts[name.lower()] = {

            "name": name,

            "title": title or m_info.get("Title") or "Executive",

            "reports_to": reports_to,

            "priority": "Relationship",  # Gold star is Relationship

            "marker": "STAR",

            "org": m_info.get("LOB") or "Gold Star Group",

            "branch": "",  # we will set branch later based on reports_to

            "products": m_info.get("Products") or "watsonx.governance / watsonx Orchestrate (AI guardrails & agent orchestration)",

            "notes": m_info.get("Notes") or "Direct Relationship contact.",

            "owner": m_info.get("Owner") or "Jerome Carlson (ATL)",

            "schedule": "Direct Relationship",

            "source": "Gold-Star",

            "verification": "Verified",

            "verification_note": "Direct Relationship contact. Verified via SAP Client Technology Strategy deck."

        }

# 2. Load POL contacts

rows_pol = list(pol_ws.iter_rows(values_only=True))[5:]

for r in rows_pol:

    if any(x is not None for x in r):

        name = r[0]

        priority = r[1]

        depth = r[2]

        branch = r[3]

        reports_to = r[4]

        marker = r[5]

        title = r[6]

        org = r[7]

        products = r[8]

        notes = r[9]

        on_list_because = r[10]

        owner = r[11]

        schedule = r[12]

        node_priority = "Low"

        if marker in ("STAR", "ORANGE"):

            node_priority = "Relationship"

        elif marker == "RED" or priority == "High":

            node_priority = "High"

        elif priority == "Medium":

            node_priority = "Medium"

        else:

            node_priority = "Low"

        contacts[name.lower()] = {

            "name": name,

            "title": title or "Staff",

            "reports_to": reports_to,

            "priority": node_priority,

            "marker": marker,

            "org": org or branch or "Other",

            "branch": branch,

            "products": products or "watsonx.governance / watsonx Orchestrate",

            "notes": notes or on_list_because or "Key technical contact.",

            "owner": owner or "Jerome Carlson (ATL)",

            "schedule": schedule or "Backlog",

            "source": "POL",

            "verification": r[13] or "Not yet run through the public-verification pass.",

            "verification_note": r[14] or "No verification notes available."

        }

# 2.7 Manually promoted to High priority per outreach strategy

for _promoted in ["jens fuchs", "madhu sridharan"]:

    if _promoted in contacts:

        contacts[_promoted]["priority"] = "High"

        contacts[_promoted]["marker"] = "RED"

# 2.6 Clean up Von Rueden duplicates & enforce correct products on his sub-org

# The POL loads "Jonathan VON RUEDEN" as a row (he was the marked target).

# We have a proper branch-head node for him already, so remove the POL copy

# to prevent a duplicate tree node.

contacts.pop("jonathan von rueden", None)

# Force watsonx Orchestrate / watsonx + High priority + Ryan Sorber on every Von Rueden sub-org contact

VON_RUEDEN_ORG = {

    "atul deo", "christian karaschewitz", "lukas lochner", "marc-oliver klein",

    "matt horenkamp", "richard grandpierre", "sophie patzelt",

    "helen oakley", "sridhar ayodhya"

}

for name_lower in VON_RUEDEN_ORG:

    if name_lower in contacts:

        contacts[name_lower]["products"] = "watsonx Orchestrate / watsonx"

        contacts[name_lower]["priority"] = "High"

        contacts[name_lower]["marker"] = "RED"

        contacts[name_lower]["owner"] = "Ryan Sorber"

# Force Concert Platform & AIOps on every Viney Khokar sub-org contact (unless already specified)

# Walk all contacts: any reporting to Viney Khokar gets Concert Platform & AIOps

for name_lower, c in contacts.items():

    rt = (c.get("reports_to") or "").lower()

    if "viney khokar" in rt or "khokar" in rt:

        if not c.get("products") or c["products"] in ("", "watsonx.governance / watsonx Orchestrate"):

            c["products"] = "Concert Platform & AIOps"

# 2.5 Add missing slide-deck contacts

extra_contacts = [

    {

        "name": "Daniel Beck",

        "title": "GM + CPO, SuccessFactors",

        "reports_to": "Manoj Swaminathan",

        "priority": "High",

        "marker": "RED",

        "org": "SuccessFactors",

        "branch": "Siva Sundaresan",

        "products": "watsonx Orchestrate (Joule A2A), Confluent (Event Streaming), watsonx.governance",

        "notes": "GM & CPO, SAP SuccessFactors. Reports to Manoj Swaminathan. Per 8/17 notes: Andy Wei (VP, Strategic Engagements) confirmed Joule is 3'4 years behind watsonx Orchestrate for A2A ' Beck owns the SuccessFactors product strategy where that gap creates IBM's entry point. Ansible AAP deployment is underway in SuccessFactors. IBM Bob is validated for Java modernization in this org. Daniel Beck is the executive sponsor needed to scale any Orchestrate or Bob deal beyond the engineering level.",

        "owner": "Eliot Frederiksen (IBM CE) / Katie McLuckie",

        "schedule": "Backlog",

        "source": "Slide-Deck",

        "verification": "Verified",

        "verification_note": "Verified from Slide 36 of Account Technology Strategy deck."

    },

    {

        "name": "Maryann Abbajay",

        "title": "CRO, SAP SuccessFactors",

        "reports_to": "Daniel Beck",

        "priority": "High",

        "marker": "RED",

        "org": "SuccessFactors",

        "branch": "Siva Sundaresan",

        "products": "watsonx Orchestrate (Joule A2A), Confluent (Event Streaming), watsonx.governance",

        "notes": "CRO, SAP SuccessFactors. Reports to Daniel Beck. Key commercial leader for SuccessFactors. Per 8/17 notes: Orchestrate A2A is 3'4 years ahead of Joule ' Abbajay's CRO role means she shapes how SuccessFactors is positioned to customers, and IBM Orchestrate directly competes with SAP Joule for AI workflow automation. Aligning with her creates a co-sell or displacement narrative at the commercial level.",

        "owner": "Josh McClure / Katie McLuckie",

        "schedule": "Backlog",

        "source": "Slide-Deck",

        "verification": "Verified",

        "verification_note": "Verified from Slide 36 of Account Technology Strategy deck."

    },

    {

        "name": "Mithun Goyal",

        "title": "DB Engineering Lead, SAP Ariba",

        "reports_to": "Mehmet Yurdal",

        "priority": "High",

        "marker": "RED",

        "org": "Ariba",

        "branch": "Lawrence Martin",

        "products": "watsonx.data Premium (Cassandra), Confluent, watsonx.governance",

        "notes": "DB Engineering Lead, SAP Ariba. Per 7/27 and 8/3 notes: Ariba is executing a Java migration from old-gen to next-gen involving 20M+ lines of code ' Mithun's DB Engineering role puts him at the center of the data layer transformation. IBM Bob's bi-directional Java modernization capability was specifically validated in Ariba conversations. watsonx.data Premium with 83% cost/time reduction for GPU-accelerated workloads is a direct fit for Cassandra migration. Top technical contact for the data-layer play in Ariba.",

        "owner": "Ryan Sorber (Prime)",

        "schedule": "Backlog",

        "source": "Slide-Deck",

        "verification": "Verified",

        "verification_note": "Verified from Slide 37 of Account Technology Strategy deck."

    },

    {

        "name": "Thomas Jackson",

        "title": "Procurement Lead, SAP Ariba",

        "reports_to": "Mehmet Yurdal",

        "priority": "High",

        "marker": "RED",

        "org": "Ariba",

        "branch": "Lawrence Martin",

        "products": "watsonx.governance, IBM Bob (App Mod)",

        "notes": "Procurement Lead, SAP Ariba. Reports to Mehmet Yurdal. Per 7/27 notes: Ariba's 20M+ line Java migration is the primary IBM Bob play ' Jackson's procurement engineering role means he evaluates and procures the tooling for this effort. IBM Bob (Code Assistant for Java) was validated as a bi-directional differentiator in Ariba conversations. watsonx.governance for procurement data lineage and compliance is also a strong fit in this org.",

        "owner": "Ryan Sorber",

        "schedule": "Backlog",

        "source": "Slide-Deck",

        "verification": "Verified",

        "verification_note": "Verified from Slide 37 of Account Technology Strategy deck."

    },

    {

        "name": "Reem Alqaysi",

        "title": "Manager, Data Engineering, SAP Ariba",

        "reports_to": "Mithun Goyal",

        "priority": "High",

        "marker": "RED",

        "org": "Ariba",

        "branch": "Lawrence Martin",

        "products": "watsonx.data Premium (Cassandra)",

        "notes": "Manager, Data Engineering, SAP Ariba. Reports to Mithun Goyal. Per 7/27 notes: Ariba's data engineering team is executing a large-scale Java and data platform migration. Alqaysi manages the data engineering function directly impacted by watsonx.data Premium (Cassandra replacement with 83% cost/time reduction). LangFlow, LangGraph, HANA vector DB, and Jupyter Notebooks are in active use ' IBM's portfolio slots naturally into this existing AI stack.",

        "owner": "Ryan Sorber",

        "schedule": "Backlog",

        "source": "Slide-Deck",

        "verification": "Verified",

        "verification_note": "Verified from Slide 37 of Account Technology Strategy deck."

    },

    {

        "name": "Santosh Tej",

        "title": "Lead Engineer, SuccessFactors",

        "reports_to": "Siva Sundaresan",

        "priority": "High",

        "marker": "RED",

        "org": "SuccessFactors",

        "branch": "Siva Sundaresan",

        "products": "watsonx Orchestrate (Joule A2A), Red Hat OpenShift / Ansible",

        "notes": "Lead Engineer, SAP SuccessFactors. Per 8/17 notes: Ansible AAP (Ansible Automation Platform) deployment is actively underway in SuccessFactors ' Tej is a key technical contact for this Red Hat OpenShift / Ansible engagement. Andy Wei confirmed Joule is 3'4 years behind Orchestrate for A2A; Tej sits in the engineering org that will eventually need to integrate or replace Joule with external AI. Josh McClure owns the Ansible thread; Ryan Sorber the data thread; Jerome is ATL coordinator.",

        "owner": "Josh McClure / Ryan Sorber / Jerome Carlson",

        "schedule": "Backlog",

        "source": "Slide-Deck",

        "verification": "Verified",

        "verification_note": "Verified from Slide 44 of Account Technology Strategy deck."

    },

    {

        "name": "Kumar Sambhav",

        "title": "Lead Engineer, SuccessFactors",

        "reports_to": "Siva Sundaresan",

        "priority": "High",

        "marker": "RED",

        "org": "SuccessFactors",

        "branch": "Siva Sundaresan",

        "products": "watsonx Orchestrate (Joule A2A), Red Hat OpenShift / Ansible",

        "notes": "Lead Engineer, SAP SuccessFactors. Per 8/17 notes: Ansible AAP deployment is underway in SuccessFactors alongside the Red Hat OpenShift rollout ' Sambhav is a co-lead engineer on this track alongside Santosh Tej. Andy Wei's Joule gap admission (3'4 years behind Orchestrate A2A) creates the strategic context for an expanded Orchestrate conversation once Ansible/OpenShift is established. Dual IBM relationship gives leverage across automation and AI layers.",

        "owner": "Josh McClure / Ryan Sorber / Jerome Carlson",

        "schedule": "Backlog",

        "source": "Slide-Deck",

        "verification": "Verified",

        "verification_note": "Verified from Slide 44 of Account Technology Strategy deck."

    },

    {

        "name": "Benjamin Blau",

        "title": "CIO Office - Internal Workloads Lead",

        "reports_to": "Sebastian Steinhaeuser",

        "priority": "High",

        "marker": "RED",

        "org": "CIO Office",

        "branch": "Strategy & Operations",

        "products": "watsonx.governance, watsonx Orchestrate, Turbonomic",

        "notes": "CIO Office ' Internal Workloads Lead. Reports to Sebastian Steinhaeuser. Per 8/10 and 8/24 notes: ESA audit issue creates ~$40M restatement risk for SAP from undeployed software ' Blau's internal workloads role means he is directly responsible for SAP running its own software. This is IBM's leverage point: IBM can help SAP accelerate internal SAP deployments to close the ESA gap. Turbonomic for SAP-on-SAP cloud cost optimization and watsonx Orchestrate for internal process automation are immediate fits. High urgency given the ESA deadline pressure.",

        "owner": "Joanne Wright (IBM CIO) / Jerome Carlson",

        "schedule": "Backlog",

        "source": "Slide-Deck",

        "verification": "Verified",

        "verification_note": "Verified from Slide 46 of Account Technology Strategy deck."

    },

    {

        "name": "Christian Freytag",

        "title": "Chief Data & Technology Officer, Corporate Processes & IT",

        "reports_to": "Sebastian Steinhaeuser",

        "priority": "High",

        "marker": "RED",

        "org": "CIO Office",

        "branch": "Strategy & Operations",

        "products": "watsonx.governance, watsonx Orchestrate",

        "notes": "Chief Data & Technology Officer, SAP Corporate Processes & IT. Reports to Sebastian Steinhaeuser. Per 8/24 notes: SAP's ESA (~$40M restatement risk from undeployed software) creates urgency for Freytag to accelerate internal platform deployments ' IBM is the partner to help SAP run its own technology at scale. As CDTO, Freytag owns the data and technology governance layer for SAP's own IT ' watsonx.governance, watsonx Orchestrate, and Turbonomic all map directly to his mandate. Strategic contact for the internal SAP-runs-SAP play.",

        "owner": "Joanne Wright (IBM CIO) / Jerome Carlson",

        "schedule": "Backlog",

        "source": "Slide-Deck",

        "verification": "Verified",

        "verification_note": "Verified from Slide 46 of Account Technology Strategy deck."

    },

    # === Von Rueden sub-org (all High priority, watsonx Orchestrate / watsonx) ===

    {

        "name": "Atul Deo",

        "title": "Development Executive, AI Product Management",

        "reports_to": "Jonathan Von Rueden",

        "priority": "High",

        "marker": "RED",

        "org": "Business AI",

        "branch": "Philipp Herzig",

        "products": "watsonx Orchestrate / watsonx",

        "notes": "Development Executive, AI Product Management. Reports to Jonathan Von Rueden. 6 direct / 57 team. Per 8/17 and 8/24 notes: Von Rueden's org owns the SAP Business AI Platform spanning BDC and BTP. Atul Deo is the execution lead for AI product management directly under Von Rueden. Andy Wei confirmed Joule (SAP's AI) is 3'4 years behind watsonx Orchestrate for A2A ' Deo's team is the product org that builds and ships Business AI, making them the primary IBM Orchestrate competitive displacement target. Eliot is pursuing Von Rueden for lunch in Palo Alto Sept 9; Deo is the next-level contact to warm simultaneously.",

        "owner": "Ryan Sorber",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Christian Karaschewitz",

        "title": "Manager, Product Management, AI Products",

        "reports_to": "Atul Deo",

        "priority": "High",

        "marker": "RED",

        "org": "Business AI",

        "branch": "Philipp Herzig",

        "products": "watsonx Orchestrate / watsonx",

        "notes": "Manager, Product Management, AI Products. Reports to Atul Deo. 7 direct / 7 team. Per 8/17 notes: Joule is 3'4 years behind watsonx Orchestrate for A2A per Andy Wei's direct statement. Karaschewitz manages the product team that defines Business AI product requirements ' making him a key target for co-development conversations and competitive displacement. IBM Orchestrate's agent framework is the alternative to Joule's lagging A2A capability. Warm alongside Deo as part of the Von Rueden org outreach.",

        "owner": "Ryan Sorber",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Lukas Lochner",

        "title": "Manager, Product Management, AI Products",

        "reports_to": "Atul Deo",

        "priority": "High",

        "marker": "RED",

        "org": "Business AI",

        "branch": "Philipp Herzig",

        "products": "watsonx Orchestrate / watsonx",

        "notes": "Manager, Product Management, AI Products. Reports to Atul Deo. 8 direct / 8 team. Per 8/17 notes: Von Rueden's Business AI org builds SAP's internal AI platform (Joule and Business AI Platform spanning BDC + BTP). Andy Wei confirmed Joule is years behind Orchestrate for A2A. Lochner manages a product team shaping the Business AI roadmap ' an IBM Orchestrate conversation at this level plants seeds for platform-level adoption. Engage as part of coordinated Von Rueden org outreach led by Ryan Sorber.",

        "owner": "Ryan Sorber",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Marc-Oliver Klein",

        "title": "Development Senior Manager, AI Products",

        "reports_to": "Atul Deo",

        "priority": "High",

        "marker": "RED",

        "org": "Business AI",

        "branch": "Philipp Herzig",

        "products": "watsonx Orchestrate / watsonx",

        "notes": "Development Senior Manager, AI Products. Reports to Atul Deo. 20 direct / 20 team. Per 8/17 notes: Von Rueden's org owns SAP Business AI Platform spanning BDC and BTP ' Klein leads the largest engineering team (20 direct) under Atul Deo, building the actual Business AI development infrastructure. IBM watsonx Orchestrate's A2A agent framework is directly competitive with what Klein's team builds. This is IBM's deepest technical entry point into the Business AI platform development org.",

        "owner": "Ryan Sorber",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Matt Horenkamp",

        "title": "Chief Product Management Expert, AI Products",

        "reports_to": "Atul Deo",

        "priority": "High",

        "marker": "RED",

        "org": "Business AI",

        "branch": "Philipp Herzig",

        "products": "watsonx Orchestrate / watsonx",

        "notes": "Chief Product Management Expert, AI Products. Reports to Atul Deo. Per 8/17 notes: Horenkamp is the senior product expert in the AI Products org under Atul Deo ' his 'Chief Expert' title means he sets the technical product standards and evaluates external technology. Andy Wei's Joule gap admission (3'4 years behind Orchestrate A2A) makes Horenkamp a prime target for an IBM Orchestrate technical evaluation conversation. He is the internal voice of authority on what Business AI products should do ' align IBM's A2A story to his evaluation criteria.",

        "owner": "Ryan Sorber",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Richard Grandpierre",

        "title": "Development Senior Manager, AI Products",

        "reports_to": "Atul Deo",

        "priority": "High",

        "marker": "RED",

        "org": "Business AI",

        "branch": "Philipp Herzig",

        "products": "watsonx Orchestrate / watsonx",

        "notes": "Development Senior Manager, AI Products. Reports to Atul Deo. 11 direct / 16 team. Per 8/17 and 8/24 notes: Grandpierre leads a development team (11 direct, 16 total) within the Business AI platform org. Von Rueden's org spans BDC and BTP ' Grandpierre's team builds components of the platform that IBM Orchestrate would integrate with or displace. Direct technical engagement at this level enables IBM to influence platform architecture decisions before they are locked in. Engage coordinated with Ryan Sorber.",

        "owner": "Ryan Sorber",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Sophie Patzelt",

        "title": "Administration Specialist, AI Products",

        "reports_to": "Atul Deo",

        "priority": "High",

        "marker": "RED",

        "org": "Business AI",

        "branch": "Philipp Herzig",

        "products": "watsonx Orchestrate / watsonx",

        "notes": "Administration Specialist, AI Products. Reports to Atul Deo. Key operational contact for scheduling and access to the Von Rueden / Atul Deo org. Per 8/24 notes: Eliot is pursuing Von Rueden for lunch in Palo Alto Sept 9 ' Patzelt as the admin specialist may be the scheduling gateway. Building a positive relationship with her accelerates access to Deo and Von Rueden. Treat as an enabler contact for the entire Business AI org outreach.",

        "owner": "Ryan Sorber",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    # === Signavio sub-org (under Andre Wenz) ===

    {

        "name": "Alessandro Paolo Manzi",

        "title": "Head of SAP Signavio Product Management",

        "reports_to": "Andre Wenz",

        "priority": "Low",

        "marker": None,

        "org": "Signavio",

        "branch": "Philipp Herzig",

        "products": "watsonx.governance / watsonx Orchestrate",

        "notes": "Head of SAP Signavio Product Management. Reports to Andre Wenz. 10 direct / 109 team. Signavio is SAP's business process intelligence and mining platform. Manzi leads the PM function (109-person team) defining Signavio's product roadmap. IBM watsonx.governance is the natural AI governance layer for process intelligence outputs. watsonx Orchestrate can automate downstream process improvements that Signavio identifies ' making IBM a natural integration partner for Signavio's next-gen AI roadmap.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Andreas Breitruck",

        "title": "Head of SAP Signavio Product & Engineering",

        "reports_to": "Andre Wenz",

        "priority": "Low",

        "marker": None,

        "org": "Signavio",

        "branch": "Philipp Herzig",

        "products": "watsonx.governance / watsonx Orchestrate",

        "notes": "Head of SAP Signavio Product & Engineering. Reports to Andre Wenz. Breitruck leads the engineering delivery for Signavio's product suite. IBM watsonx.governance for AI model monitoring within process mining outputs is a natural fit. IBM Bob (Code Assistant for Java) is relevant if Signavio has Java modernization needs in its engineering stack. Engage as part of the Signavio org outreach through Eliot Frederiksen.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Bastian Steinert",

        "title": "Head of Signavio Technology & Architecture",

        "reports_to": "Andre Wenz",

        "priority": "Low",

        "marker": None,

        "org": "Signavio",

        "branch": "Philipp Herzig",

        "products": "watsonx.governance / watsonx Orchestrate",

        "notes": "Head of Signavio Technology & Architecture. Reports to Andre Wenz. 5 direct / 15 team. Steinert owns the technology and architecture layer for SAP Signavio ' making him the key decision-maker for any IBM platform integration into Signavio's architecture. IBM watsonx.governance and watsonx Orchestrate integrations into the Signavio process intelligence stack would go through Steinert's architecture review. Critical technical contact for the Signavio platform play.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Gregor Berg",

        "title": "Head of SAP Signavio.NEXT",

        "reports_to": "Andre Wenz",

        "priority": "Low",

        "marker": None,

        "org": "Signavio",

        "branch": "Philipp Herzig",

        "products": "watsonx.governance / watsonx Orchestrate",

        "notes": "Head of SAP Signavio.NEXT. Reports to Andre Wenz. 15 direct / 25 team. Signavio.NEXT is the next-generation product line for SAP's process intelligence platform. Berg leads this innovation team ' making him a forward-looking target for IBM's next-gen AI governance and orchestration story. IBM watsonx Orchestrate's A2A agent framework is a natural fit for automating the process improvement recommendations that Signavio.NEXT will generate. High-value contact for the innovation/roadmap conversation.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Lukas N.P. Egger",

        "title": "Head of SAP Signavio Product Innovation Office",

        "reports_to": "Andre Wenz",

        "priority": "Low",

        "marker": None,

        "org": "Signavio",

        "branch": "Philipp Herzig",

        "products": "watsonx.governance / watsonx Orchestrate",

        "notes": "Head of SAP Signavio Product Innovation Office. Reports to Andre Wenz. 21 direct / 21 team. Egger leads the Signavio Product Innovation Office ' the team that defines Signavio's future product direction. With 21 direct reports, this is the largest single team under Andre Wenz. IBM watsonx Orchestrate and watsonx.governance are natural platform companions for Signavio's innovation roadmap. This is the entry point for positioning IBM as Signavio's AI infrastructure partner of choice.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Rebekka Kotinis",

        "title": "Head of SAP Signavio P&E Experience",

        "reports_to": "Andre Wenz",

        "priority": "Low",

        "marker": None,

        "org": "Signavio",

        "branch": "Philipp Herzig",

        "products": "watsonx.governance / watsonx Orchestrate",

        "notes": "Head of SAP Signavio P&E Experience. Reports to Andre Wenz. 10 direct / 91 team. Kotinis leads the P&E Experience function (91 total team) for Signavio ' covering the developer and user experience layer. IBM Bob (Code Assistant) is highly relevant for improving developer experience and productivity within Signavio's engineering org. IBM watsonx Orchestrate for workflow automation in the user experience layer is also a strong fit. Large team size makes this a high-leverage contact.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Torsten Zube",

        "title": "Head of SAP Signavio P&E Engineering",

        "reports_to": "Andre Wenz",

        "priority": "Low",

        "marker": None,

        "org": "Signavio",

        "branch": "Philipp Herzig",

        "products": "watsonx.governance / watsonx Orchestrate",

        "notes": "Head of SAP Signavio P&E Engineering. Reports to Andre Wenz. 10 direct / 678 team. Zube leads the largest engineering org under Andre Wenz (678 total) ' the full Signavio product engineering team. At 678 engineers, IBM Bob for developer productivity (code acceleration) is the highest-ROI play. IBM watsonx.governance for AI compliance in engineering workflows and IBM Turbonomic for cloud cost optimization of the engineering infrastructure are also relevant. Highest team-size contact in the Signavio sub-org.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    # === LeanIX sub-org (under Dominik Rose ' Steffen Wittmann) ===

    {

        "name": "Steffen Wittmann",

        "title": "Head of SAP LeanIX Engineering",

        "reports_to": "Dominik Rose",

        "priority": "Low",

        "marker": None,

        "org": "LeanIX",

        "branch": "Philipp Herzig",

        "products": "watsonx.governance / watsonx Orchestrate",

        "notes": "Head of SAP LeanIX Engineering. Reports to Dominik Rose. 6 direct / 337 team. Wittmann leads the entire LeanIX engineering organization (337 engineers). LeanIX is SAP's Enterprise Architecture Management platform ' IBM watsonx.governance for AI model governance in EA tooling is a direct fit. IBM Bob for developer productivity across a 337-engineer org is a high-ROI play. Turbonomic for cloud cost optimization of LeanIX's engineering infrastructure is also relevant. Top engineering contact for the LeanIX platform plays.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Arjun Kohli",

        "title": "Head of Engineering Information Security, LeanIX",

        "reports_to": "Steffen Wittmann",

        "priority": "Low",

        "marker": None,

        "org": "LeanIX",

        "branch": "Philipp Herzig",

        "products": "watsonx.governance / watsonx Orchestrate",

        "notes": "Head of Engineering Information Security. Reports to Steffen Wittmann. 5 direct / 5 team. Kohli leads information security for LeanIX engineering. Per 8/10 notes: AI security is Herzig's top priority (per Tanya Burley) ' this mindset flows down to product engineering security leads like Kohli. IBM watsonx.governance for AI security guardrails and Guardium for data security compliance are the primary plays. Key contact for security-angle IBM conversations within LeanIX engineering.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Jonas Rathert",

        "title": "PM AI Acceleration, LeanIX",

        "reports_to": "Steffen Wittmann",

        "priority": "Low",

        "marker": None,

        "org": "LeanIX",

        "branch": "Philipp Herzig",

        "products": "watsonx.governance / watsonx Orchestrate",

        "notes": "PM AI Acceleration. Reports to Steffen Wittmann. 4 direct / 4 team. Rathert's title 'AI Acceleration' makes him the most directly relevant LeanIX contact for IBM's AI portfolio. He defines and drives AI integration into LeanIX's engineering workflows ' IBM watsonx Orchestrate for AI-powered EA automation and IBM Bob for AI-accelerated development are precisely what Rathert's role is evaluating. Highest-priority technical contact in the LeanIX sub-org for an AI conversation.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Jost Novljan",

        "title": "Head of Engineering Capabilities, LeanIX",

        "reports_to": "Steffen Wittmann",

        "priority": "Low",

        "marker": None,

        "org": "LeanIX",

        "branch": "Philipp Herzig",

        "products": "watsonx.governance / watsonx Orchestrate",

        "notes": "Head of Engineering Capabilities. Reports to Steffen Wittmann. 6 direct / 222 team. Novljan leads engineering capabilities for LeanIX (222-person team) ' covering the platform enablement layer. IBM Bob for developer productivity at scale (222 engineers) is the highest-ROI play. IBM watsonx.governance for AI compliance in engineering capability tooling and Turbonomic for infrastructure cost optimization are also relevant. Second-largest team under Wittmann.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Konrad Walzer",

        "title": "Head of Engineering DevX & AI Acceleration, LeanIX",

        "reports_to": "Steffen Wittmann",

        "priority": "Low",

        "marker": None,

        "org": "LeanIX",

        "branch": "Philipp Herzig",

        "products": "watsonx.governance / watsonx Orchestrate",

        "notes": "Head of Engineering DevX & AI Acceleration. Reports to Steffen Wittmann. 7 direct / 23 team. Walzer leads Developer Experience and AI Acceleration for LeanIX engineering ' his title is the most IBM-aligned role in the LeanIX sub-org. IBM Bob (Code Assistant) for developer experience improvement and IBM watsonx Orchestrate for AI-accelerated engineering workflows are directly what Walzer is tasked with delivering. High-value target for a technical AI product conversation.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Nandor Orvos",

        "title": "Head of Engineering CSE & Support, LeanIX",

        "reports_to": "Steffen Wittmann",

        "priority": "Low",

        "marker": None,

        "org": "LeanIX",

        "branch": "Philipp Herzig",

        "products": "watsonx.governance / watsonx Orchestrate",

        "notes": "Head of Engineering CSE & Support. Reports to Steffen Wittmann. 4 direct / 46 team. Orvos leads Customer Success Engineering and Support for LeanIX. IBM Turbonomic for automated cloud cost and performance optimization reduces the support burden ' a direct pain-point play. IBM watsonx Orchestrate for automated support workflow routing is also relevant for a CSE org. Engage as part of the broader LeanIX engineering outreach through Wittmann.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Roman Saul",

        "title": "Head of Engineering Shared Services, LeanIX",

        "reports_to": "Steffen Wittmann",

        "priority": "Low",

        "marker": None,

        "org": "LeanIX",

        "branch": "Philipp Herzig",

        "products": "watsonx.governance / watsonx Orchestrate",

        "notes": "Head of Engineering Shared Services. Reports to Steffen Wittmann. 16 direct / 31 team. Saul leads shared engineering services for LeanIX ' the cross-cutting platform functions (tooling, CI/CD, infra). IBM Bob for developer productivity in shared services, IBM Turbonomic for shared infrastructure cost optimization, and IBM watsonx.governance for cross-org AI compliance tooling are all relevant plays. 16 direct reports means Saul has significant team influence ' engage as part of the Wittmann-led LeanIX outreach.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    # === BDC sub-org (under Irfan Khan ' Prakash Nanduri ' Milinda Vitharana) ===

    {

        "name": "Prakash Nanduri",

        "title": "Head of BDC&I Product Management",

        "reports_to": "Irfan Khan",

        "priority": "Low",

        "marker": None,

        "org": "Business Data Cloud",

        "branch": "Philipp Herzig",

        "products": "watsonx.data Premium / watsonx.governance / Confluent",

        "notes": "Head of BDC&I Product Management. Reports to Irfan Khan. 9 direct / 174 team. Per 7/20 notes: BDC is the master data team for all of SAP ' Nanduri is the product management lead for the entire BDC&I organization (174-person team). IBM watsonx.data Premium with GPU acceleration delivered 83% cost/time reduction in BDC conversations. Confluent has an OEM partnership with SAP Datasphere (Nanduri's product portfolio). Top product management contact for the watsonx.data Premium and Confluent plays in the BDC org.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Milinda Vitharana",

        "title": "Head of BDC&I FOS Product Management",

        "reports_to": "Prakash Nanduri",

        "priority": "Low",

        "marker": None,

        "org": "Business Data Cloud",

        "branch": "Philipp Herzig",

        "products": "watsonx.data Premium / watsonx.governance / Confluent",

        "notes": "Head of BDC&I FOS Product Management. Reports to Prakash Nanduri. 7 direct / 16 team. Per 7/20 notes: BDC is SAP's master data team ' Vitharana leads the FOS (Foundation & Operations Services) PM team within BDC. IBM watsonx.data Premium's 83% GPU acceleration resonated strongly in BDC discussions. She sits between Nanduri and the engineering leads (Horne, Erattemparambil) ' a key product decision influencer for any watsonx.data Premium or Confluent deployment in BDC.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Robert Horne",

        "title": "Head of BDC&I P&A PM Provision, Monitor & Adminis",

        "reports_to": "Milinda Vitharana",

        "priority": "Low",

        "marker": None,

        "org": "Business Data Cloud",

        "branch": "Philipp Herzig",

        "products": "watsonx.data Premium / watsonx.governance / Confluent",

        "notes": "Head of BDC&I P&A PM Provision, Monitor & Administration. Reports to Milinda Vitharana. 9 direct / 9 team. Per 7/20 notes: BDC is the master data team for all of SAP ' Horne's Provision, Monitor & Administration role means he directly oversees the operational tooling stack that IBM watsonx.data Premium would enter. 83% cost/time reduction with GPU acceleration is the key proof point. Horne controls the provisioning and monitoring infrastructure where IBM replaces or augments existing data tools.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    {

        "name": "Sreedevi Erattemparambil",

        "title": "BDC&I P&A PM Provision, Monitor & Adminis",

        "reports_to": "Milinda Vitharana",

        "priority": "Low",

        "marker": None,

        "org": "Business Data Cloud",

        "branch": "Philipp Herzig",

        "products": "watsonx.data Premium / watsonx.governance / Confluent",

        "notes": "BDC&I P&A PM Provision, Monitor & Administration. Reports to Milinda Vitharana. Per 7/20 notes: within the BDC master data team where IBM watsonx.data Premium (83% GPU cost/time reduction) resonated. Erattemparambil works directly in the provisioning and administration PM function ' a useful technical contact for understanding BDC's current tooling gaps and positioning watsonx.data Premium as the replacement. Engage alongside Horne and Vitharana in the BDC data platform conversations.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Backlog",

        "source": "Org Chart Image",

        "verification": "Verified",

        "verification_note": "Verified from SAP org chart image."

    },

    # === Viney Khokar org (Concert Platform & AIOps) ===

    {

        "name": "Viney Khokar",

        "title": "VP, IT Operations & AIOps, SAP",

        "reports_to": "Philipp Herzig",

        "priority": "High",

        "marker": "RED",

        "org": "IT Operations & AIOps",

        "branch": "Philipp Herzig",

        "products": "Concert Platform & AIOps",

        "notes": "VP, IT Operations & AIOps at SAP. Reports to Philipp Herzig. IBM Concert Platform is the primary play ' Concert provides the unified AIOps and IT operations observability layer that directly addresses SAP's operational complexity at scale. Key target for IBM's Concert Platform expansion into SAP's internal IT operations.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Week 1",

        "source": "Org Chart Image",

        "verification": "Not yet verified",

        "verification_note": "Added per outreach strategy. Verify title and reporting line."

    },

    # === Email-archive confirmed contacts (not in Excel) ===

    {

        "name": "Chuck Firkin",

        "title": "VP, Sovereign Cloud Technology & Engineering",

        "reports_to": "Martin Merz",

        "priority": "Relationship",

        "marker": "STAR",

        "org": "Sovereign Cloud",

        "branch": "Martin Merz",

        "products": "Guardium QSE & QSR, watsonx.governance, IBM Bob",

        "notes": "VP, Sovereign Cloud Technology & Engineering. Reports to Martin Merz. Sits on Sovereignty Board. Per 8/10 notes: AI security is top of Herzig's mind (per Tanya Burley, Herzig's chief of staff) ' Firkin owns the infrastructure layer where that lands. Discussed IBM Lightwell for open-source support and IBM StorageCEPH as cost-reduction plays for sovereign cloud. Direct relationship ' high-value gate to the Sovereign Cloud platform stack. Email confirmed direct relationship with Eliot Frederiksen (IBM Meeting Follow Up 22.07.26).",

        "owner": "Eliot Frederiksen (IBM Client Executive)",

        "schedule": "N/A",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via IBM Meeting Follow Up email thread (Eliot to Chuck, Jerome CC'd)."

    },

    {

        "name": "Gokul Naidu",

        "title": "VP Engineering, SAP SuccessFactors",

        "reports_to": "Siva Sundaresan",

        "priority": "High",

        "marker": "RED",

        "org": "SuccessFactors",

        "branch": "Siva Sundaresan",

        "products": "watsonx Orchestrate (Joule A2A), IBM Bob, watsonx.governance",

        "notes": "VP Engineering, SAP SuccessFactors. Reports to Siva Sundaresan. Per email: Josh McClure and Eliot Frederiksen both engaged Gokul (Save the Date for IBM Leadership Exchange). Joule A2A gap is the primary play in SuccessFactors engineering. IBM Bob for Java modernization validated in this org. A key technical executive for the SuccessFactors IBM play.",

        "owner": "Eliot Frederiksen (IBM Client Executive)",

        "schedule": "Week 1",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via Save the Date email thread (Josh to Gokul, Eliot CC'd)."

    },

    {

        "name": "Nick Totten",

        "title": "Senior Director, SAP Engineering",

        "reports_to": "Balaji Balasubramanian",

        "priority": "High",

        "marker": "RED",

        "org": "Customer Experience (CX)",

        "branch": "Balaji Balasubramanian",

        "products": "watsonx Orchestrate / watsonx.governance",

        "notes": "Senior Director, SAP Engineering. Reports to Balaji Balasubramanian. Jerome Carlson conducted outreach directly (Eliot CC'd). Key technical contact in the CX engineering org for watsonx Orchestrate automation and governance plays.",

        "owner": "Eliot Frederiksen (IBM Client Executive)",

        "schedule": "Week 2",

        "source": "Email Archive",

        "verification": "Not yet verified",

        "verification_note": "Confirmed via email outreach thread (Jerome to Nick, Eliot CC'd)."

    },

    {

        "name": "Collin Clark",

        "title": "Engineering Director, SAP Concur",

        "reports_to": "Balaji Balasubramanian",

        "priority": "Medium",

        "marker": None,

        "org": "Concur",

        "branch": "Balaji Balasubramanian",

        "products": "watsonx Orchestrate / IBM Bob / watsonx.governance",

        "notes": "Engineering Director, SAP Concur. Reports to Balaji Balasubramanian. Jerome Carlson led outreach (Eliot CC'd). Concur's expense and travel automation stack is a natural fit for IBM watsonx Orchestrate. IBM Bob relevant for Java modernization in Concur engineering.",

        "owner": "Eliot Frederiksen (IBM Client Executive)",

        "schedule": "Week 6",

        "source": "Email Archive",

        "verification": "Not yet verified",

        "verification_note": "Confirmed via email outreach thread (Jerome to Collin, Eliot CC'd)."

    },

    {

        "name": "Carey Main",

        "title": "VP, Security Architecture & Engineering, SAP Sovereign Cloud",

        "reports_to": "Chuck Firkin",

        "priority": "High",

        "marker": "RED",

        "org": "Sovereign Cloud",

        "branch": "Martin Merz",

        "products": "Guardium QSE & QSR, watsonx.governance, IBM Bob",

        "notes": "VP, Security Architecture & Engineering, SAP Sovereign Cloud. Reports through Chuck Firkin. Josh McClure is primary contact (Josh primary, Jerome CC'd). Per email threads: Sovereign Cloud security is a live engagement area ' Guardium for log management and watsonx.governance for AI security are the plays. Key technical decision-maker in the Sovereign Cloud security stack.",

        "owner": "Josh McClure",

        "schedule": "Week 1",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via email thread (Josh primary, Jerome CC'd)."

    },

    {

        "name": "Brian Morrison",

        "title": "VP, Technology & Engineering, SAP NS2",

        "reports_to": "Harish Luthra",

        "priority": "High",

        "marker": "RED",

        "org": "NS2",

        "branch": "Martin Merz",

        "products": "Guardium, watsonx.governance, watsonx.data Premium, IBM Bob",

        "notes": "VP, Technology & Engineering, SAP NS2. Reports to Harish Luthra. Ryan Sorber is primary contact (Ryan sent Quick Sync directly to Brian, no other IBM sellers). NS2 is replacing Splunk ' Guardium delivers 50'90% cost savings and is FedRAMP authorized. Data logging modernization is a live pain point. Key technical executive for any NS2 platform or security tooling deal.",

        "owner": "Ryan Sorber",

        "schedule": "Week 2",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via Quick Sync email thread (Ryan to Brian, sole IBM sender)."

    },

    {

        "name": "Samantha Combs",

        "title": "Director, Engineering, SAP NS2",

        "reports_to": "Brian Morrison",

        "priority": "High",

        "marker": "RED",

        "org": "NS2",

        "branch": "Martin Merz",

        "products": "Guardium, watsonx.governance, IBM Bob",

        "notes": "Director, Engineering, SAP NS2. Reports to Brian Morrison. Ryan Sorber is primary contact (Ryan sent Sam & Ryan Sync directly to Samantha). NS2 engineering is actively evaluating Splunk replacements ' Guardium is the lead play. IBM Bob for code modernization is relevant in the NS2 engineering stack.",

        "owner": "Ryan Sorber",

        "schedule": "Week 3",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via Sam & Ryan Sync email thread (Ryan to Sam, sole IBM sender)."

    },

    {

        "name": "Jim Rotan",

        "title": "Director, Security Engineering, SAP",

        "reports_to": "Martin Merz",

        "priority": "High",

        "marker": "RED",

        "org": "Sovereign Cloud & NS2",

        "branch": "Martin Merz",

        "products": "Guardium QSE & QSR, watsonx.governance",

        "notes": "Director, Security Engineering, SAP. Josh McClure is primary contact (Josh primary, Jerome CC'd). Active security engineering engagement ' Guardium for log management and watsonx.governance for AI security compliance are the primary plays. Key technical contact for the SAP security org.",

        "owner": "Josh McClure",

        "schedule": "Week 2",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via email thread (Josh primary, Jerome CC'd)."

    },

    {

        "name": "John Jiang",

        "title": "Director, IAM & Identity Governance, SAP",

        "reports_to": "Martin Merz",

        "priority": "High",

        "marker": "RED",

        "org": "Sovereign Cloud & NS2",

        "branch": "Martin Merz",

        "products": "IBM Verify / watsonx.governance / Guardium",

        "notes": "Director, IAM & Identity Governance, SAP. Josh McClure is primary contact (Josh primary, Jerome CC'd). IBM Verify (IAM) and Guardium (privileged access monitoring) are the primary plays. Active IAM governance engagement underway.",

        "owner": "Josh McClure",

        "schedule": "Week 3",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via email thread (Josh primary, Jerome CC'd)."

    },

    {

        "name": "James Reidy",

        "title": "Senior Manager, IAM Engineering, SAP",

        "reports_to": "John Jiang",

        "priority": "High",

        "marker": "RED",

        "org": "Sovereign Cloud & NS2",

        "branch": "Martin Merz",

        "products": "IBM Verify / Guardium",

        "notes": "Senior Manager, IAM Engineering, SAP. Reports to John Jiang. Josh McClure is primary contact (Josh primary, Jerome CC'd). Deep technical IAM contact ' IBM Verify for identity lifecycle and Guardium for privileged access monitoring are the active plays in this org.",

        "owner": "Josh McClure",

        "schedule": "Week 4",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via email thread (Josh primary, Jerome CC'd)."

    },

    {

        "name": "Jackson Borges",

        "title": "Principal Engineer, SAP Spend Management / Ariba",

        "reports_to": "Mehmet Yurdal",

        "priority": "Medium",

        "marker": None,

        "org": "Ariba / Spend Management",

        "branch": "Lawrence Martin",

        "products": "IBM Bob (Code Assistant for Java) / watsonx.governance",

        "notes": "Principal Engineer, SAP Spend Management / Ariba. Reports to Mehmet Yurdal. Josh McClure is primary contact (Josh primary, Jerome CC'd). Ariba's 20M+ line Java migration is the primary IBM Bob play. watsonx.governance for procurement data lineage and compliance is also a strong fit in this org.",

        "owner": "Josh McClure",

        "schedule": "Week 6",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via email thread (Josh primary, Jerome CC'd)."

    },

    {

        "name": "Mark DeVita",

        "title": "Director, Network Engineering, SAP",

        "reports_to": "Martin Merz",

        "priority": "Medium",

        "marker": None,

        "org": "Sovereign Cloud & NS2",

        "branch": "Martin Merz",

        "products": "Concert Platform & AIOps / Turbonomic",

        "notes": "Director, Network Engineering, SAP. Josh McClure is primary contact (Josh primary, Jerome CC'd). Network engineering is a Concert Platform & AIOps target ' unified observability across network, storage, and compute is the play. IBM Turbonomic for network resource optimization is also relevant.",

        "owner": "Josh McClure",

        "schedule": "Week 7",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via email thread (Josh primary, Jerome CC'd)."

    },

    {

        "name": "Adam Levinson",

        "title": "Director, CDX Platform Engineering, SAP",

        "reports_to": "Mona Mohan George",

        "priority": "High",

        "marker": "RED",

        "org": "CDX",

        "branch": "Vijay Seethapathy",

        "products": "Concert Platform & AIOps / Turbonomic / watsonx Orchestrate",

        "notes": "Director, CDX Platform Engineering, SAP. Reports to Mona Mohan George. Josh McClure is primary contact (Josh primary, Jerome CC'd). CDX is the SAP internal engineering platform ' the AIOps POC budget was approved through this org. Concert Platform & AIOps is the primary play. IBM Turbonomic for CDX cloud cost optimization and watsonx Orchestrate for platform automation are also relevant.",

        "owner": "Josh McClure",

        "schedule": "Week 1",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via email thread (Josh primary, Jerome CC'd)."

    },

    {

        "name": "Konstantinos Michalitsis",

        "title": "Senior Director, CDX Automation Engineering, SAP",

        "reports_to": "Mona Mohan George",

        "priority": "High",

        "marker": "RED",

        "org": "CDX",

        "branch": "Vijay Seethapathy",

        "products": "Concert Platform & AIOps / watsonx Orchestrate / Ansible AAP",

        "notes": "Senior Director, CDX Automation Engineering, SAP. Reports to Mona Mohan George. Jerome Carlson is primary contact (introduced via Aniya Govan). CDX Automation is the direct target for Ansible AAP deployment and watsonx Orchestrate automation. Concert Platform & AIOps for unified observability across the CDX automation stack is the active play.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Week 2",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via email intro thread (Aniya Govan to Jerome)."

    },

    {

        "name": "Anthony Sanchez",

        "title": "VP, AI & Data Observability, SAP NS2",

        "reports_to": "Harish Luthra",

        "priority": "High",

        "marker": "RED",

        "org": "NS2",

        "branch": "Martin Merz",

        "products": "Concert Platform & AIOps / watsonx.governance / Guardium",

        "notes": "VP, AI & Data Observability, SAP NS2. Reports to Harish Luthra. Jerome Carlson is primary contact (introduced via Aniya Govan). NS2 AI & Data Observability is a direct Concert Platform target ' unified AI observability in a FedRAMP environment. watsonx.governance for AI compliance and Guardium for data security are co-play products.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Week 1",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via email intro thread (Aniya Govan to Jerome)."

    },

    {

        "name": "Harsha Ravuri",

        "title": "Director, Storage & Infrastructure, SAP",

        "reports_to": "Martin Merz",

        "priority": "Medium",

        "marker": None,

        "org": "Sovereign Cloud & NS2",

        "branch": "Martin Merz",

        "products": "IBM Storage CEPH / Turbonomic / Concert Platform",

        "notes": "Director, Storage & Infrastructure, SAP. Reports to Martin Merz. Jerome Carlson is primary contact (Jerome sent outreach, Candace Dowling CC'd). IBM StorageCEPH is the primary play for Sovereign Cloud storage cost reduction. Turbonomic for infrastructure cost optimization and Concert Platform for storage observability are co-plays. Active engagement underway.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Week 5",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via email outreach thread (Jerome to Harsha, Candace CC'd)."

    },

    {

        "name": "Ryan Duncan",

        "title": "Director, Data Center Management, SAP",

        "reports_to": "Martin Merz",

        "priority": "Medium",

        "marker": None,

        "org": "Sovereign Cloud & NS2",

        "branch": "Martin Merz",

        "products": "IBM Storage CEPH / Turbonomic / Concert Platform",

        "notes": "Director, Data Center Management, SAP. Josh McClure is primary contact (Josh primary, Jerome CC'd). Data center management is a key Concert Platform & AIOps target for unified infrastructure observability. IBM StorageCEPH for cost reduction and Turbonomic for workload optimization are the primary plays.",

        "owner": "Josh McClure",

        "schedule": "Week 6",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via email thread (Josh primary, Jerome CC'd)."

    },

    {

        "name": "Gana Sadasivam",

        "title": "VP Engineering, SAP Ariba",

        "reports_to": "Mehmet Yurdal",

        "priority": "High",

        "marker": "RED",

        "org": "Ariba / Spend Management",

        "branch": "Lawrence Martin",

        "products": "Concert Platform & AIOps / IBM Bob (Code Assistant) / watsonx.governance",

        "notes": "VP Engineering, SAP Ariba. Reports to Mehmet Yurdal. Jerome Carlson is primary contact (Gana made the introductions to Dan Cooley and Mehmet Yurdal on Jerome's behalf). Ariba's 20M+ line Java migration is a direct IBM Bob play. Concert Platform & AIOps for Ariba's microservice observability and OpenTelemetry migration. watsonx.governance for procurement AI compliance. Key connector in the Ariba org ' Gana has directly opened doors for IBM.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Week 1",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via Intro: Jerome, Dan and Intro: Jerome, Mehmet email threads."

    },

    {

        "name": "John Mazzolini",

        "title": "Director, Cloud Operations, SAP CDX",

        "reports_to": "Mona Mohan George",

        "priority": "High",

        "marker": "RED",

        "org": "CDX",

        "branch": "Vijay Seethapathy",

        "products": "Concert Platform & AIOps / Turbonomic / watsonx Orchestrate",

        "notes": "Director, Cloud Operations, SAP CDX. Reports to Mona Mohan George. Josh McClure is primary contact (Josh ran IBM Updates thread directly with Mona and Mazzolini). CDX Cloud Operations is the operational hub for the AIOps POC ' Concert Platform & AIOps is the primary play. IBM Turbonomic for cloud cost optimization in CDX operations and watsonx Orchestrate for automated operations are co-plays.",

        "owner": "Josh McClure",

        "schedule": "Week 2",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via IBM Updates email thread (Josh to Mona and Mazzolini)."

    },

    {

        "name": "Ajit Kaicker",

        "title": "VP, ESA & Strategic Partnerships, SAP",

        "reports_to": "Philipp Herzig",

        "priority": "High",

        "marker": "RED",

        "org": "CTO Org / Business AI",

        "branch": "Philipp Herzig",

        "products": "Enterprise Solutions Agreement (ESA) / watsonx.governance / watsonx Orchestrate",

        "notes": "VP, ESA & Strategic Partnerships, SAP. Reports to Philipp Herzig. Eliot Frederiksen and Jerome Carlson are both on the ESA AI priorities thread. Ajit owns the ESA from the SAP side ' the ~$40M restatement risk from undeployed IBM software under the ESA makes this a critical relationship. Eliot is the senior IBM contact for this conversation.",

        "owner": "Eliot Frederiksen (IBM Client Executive)",

        "schedule": "Week 1",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via ESA AI priorities email thread (Eliot + Jerome on thread)."

    },

    {

        "name": "Daniel Cooley",

        "title": "Director, Platform Engineering, SAP Ariba",

        "reports_to": "Gana Sadasivam",

        "priority": "High",

        "marker": "RED",

        "org": "Ariba / Spend Management",

        "branch": "Lawrence Martin",

        "products": "Concert Platform & AIOps / IBM Bob / watsonx.governance",

        "notes": "Director, Platform Engineering, SAP Ariba. Reports to Gana Sadasivam. Jerome Carlson is primary contact (Gana introduced Jerome to Dan). Ariba's microservice telemetry and OpenTelemetry adoption is the Concert Platform entry point ' discussed inter-microservice traceability, correlation IDs, cross-LOB debugging across separate Dynatrace tenants, and Kafka visibility in the next-gen Ariba stack. IBM Bob for the 20M+ line Java migration is a strong co-play.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Week 2",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via Intro: Jerome, Dan email thread (Gana to Dan and Jerome)."

    },

    {

        "name": "Gregor Tielsch",

        "title": "CTO, SAP Ariba",

        "reports_to": "Mehmet Yurdal",

        "priority": "High",

        "marker": "RED",

        "org": "Ariba / Spend Management",

        "branch": "Lawrence Martin",

        "products": "IBM Bob (Code Assistant) / Concert Platform & AIOps / watsonx.governance",

        "notes": "CTO, SAP Ariba. Reports to Mehmet Yurdal. Dirk Basenach met directly with Gregor to discuss AI tooling in Ariba's CI/CD pipeline; Jerome CC'd. Ariba has already invested in Claude and Cursor for CI/CD ' IBM Bob differentiators (inference cost, open-source LLM routing) are the competitive angle. IBM Bob's bi-directional Java migration capability is the key proof point for the 20M+ line Ariba Java codebase.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Week 3",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via Virusscanner and Bob Differentiators email thread (Dirk to Gregor, Jerome CC'd)."

    },

    {

        "name": "Mehmet Yurdal",

        "title": "SVP, Head of Ariba Engineering",

        "reports_to": "Lawrence Martin",

        "priority": "High",

        "marker": "RED",

        "org": "Ariba / Spend Management",

        "branch": "Lawrence Martin",

        "products": "IBM Bob (Code Assistant) / Concert Platform & AIOps / watsonx.governance",

        "notes": "SVP, Head of Ariba Engineering. Reports to Lawrence Martin. Jerome Carlson is primary contact (Gana Sadasivam introduced Jerome to Mehmet ' described as the busiest person in Ariba who transformed Ariba Engineering and delivered Nextgen Ariba). Ariba's 20M+ line Java migration is the primary IBM Bob play. Concert Platform for Ariba microservice observability. Mehmet is the top engineering executive for the entire Ariba org ' the key decision-maker for any IBM tooling deal in Ariba.",

        "owner": "Jerome Carlson (ATL)",

        "schedule": "Week 1",

        "source": "Email Archive",

        "verification": "Verified",

        "verification_note": "Confirmed via Intro: Jerome, Mehmet email thread (Gana to Mehmet and Jerome)."

    },

]

for ec in extra_contacts:

    contacts[ec["name"].lower()] = ec

# Trace branches for Gold Star contacts

for k, c in contacts.items():

    if c["source"] == "Gold-Star":

        curr = c

        visited = set()

        branch_name = ""

        while curr and curr.get("reports_to"):

            mgr = curr["reports_to"].lower()

            if mgr in visited:

                break

            visited.add(mgr)

            if mgr in contacts:

                curr = contacts[mgr]

                if curr.get("branch"):

                    branch_name = curr["branch"]

                    break

            else:

                for bh in ["lawrence martin", "balaji balasubramanian", "siva sundaresan", "vijay seethapathy", "martin merz", "cedric bru"]:

                    if bh in mgr:

                        branch_name = bh.title()

                        break

                break

        if branch_name:

            c["branch"] = branch_name

            if c["org"] == "Gold Star Group" or not c["org"]:

                c["org"] = branch_name

# Now let's build the hierarchical tree!

# Create standard branch head nodes:

branch_heads = {

    "lawrence martin": {

        "name": "Lawrence Martin",

        "title": "CPO & Head of Public Cloud Engineering",

        "org": "Public Cloud Engineering",

        "priority": "Low",

        "marker": None,

        "notes": "CPO & Head of Public Cloud Engineering. Reports to Manoj Swaminathan. Owns the public cloud engineering platform ' IBM watsonx.governance and watsonx Orchestrate are the AI governance and automation layers for public cloud at scale. IBM Bob (Code Assistant) is relevant for the large Java modernization footprint in public cloud engineering.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "balaji balasubramanian": {

        "name": "Balaji Balasubramanian",

        "title": "SVP & Head of Customer Experience (CX)",

        "org": "Customer Experience (CX)",

        "priority": "Low",

        "marker": None,

        "notes": "SVP & Head of Customer Experience (CX). Reports to Manoj Swaminathan. CX engineering drives product adoption at scale ' IBM watsonx Orchestrate for AI-powered customer engagement workflows and watsonx.governance for responsible AI in customer-facing products are natural fits.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "siva sundaresan": {

        "name": "Siva Sundaresan",

        "title": "SVP & Head of SuccessFactors HCM Engineering",

        "org": "SuccessFactors HCM Engineering",

        "priority": "Low",

        "marker": None,

        "notes": "SVP & Head of SuccessFactors HCM Engineering. Reports to Manoj Swaminathan. Per 8/17 notes: Andy Wei (in Sundaresan's extended org) confirmed Joule is 3'4 years behind watsonx Orchestrate for A2A ' Sundaresan's engineering org is where that competitive gap plays out. Ansible AAP deployment is underway in SuccessFactors. IBM Bob for Java modernization is validated. Top-of-org contact for the entire SuccessFactors IBM play.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "vijay seethapathy": {

        "name": "Vijay Seethapathy",

        "title": "Global Head of Cloud ERP",

        "org": "Cloud ERP",

        "priority": "High",

        "marker": "RED",

        "notes": "Global Head of Cloud ERP. Reports to Manoj Swaminathan. Per 7/20 notes: Mona Mohan George and CDX report up through Vijay ' the approved AIOps POC budget and the Deo Caruana SRE decision-making chain all ultimately flow to Seethapathy. Promoted to High priority. IBM Turbonomic for Cloud ERP cost optimization and watsonx Orchestrate for ERP workflow automation are the primary plays. Critical contact for the entire Cloud ERP and CDX IBM engagement.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "martin merz": {

        "name": "Martin Merz",

        "title": "SVP & Head of Sovereign Cloud Technology & Engineering / NS2",

        "org": "Sovereign Cloud & NS2",

        "priority": "Low",

        "marker": None,

        "notes": "SVP & Head of Sovereign Cloud Technology & Engineering / NS2. Reports to Manoj Swaminathan. Per 8/3 and 8/10 notes: Sovereign Cloud has active pain around open-source software costs (IBM Lightwell), storage (IBM StorageCEPH), and AI security (top of Herzig's mind). NS2 (reporting through Merz) is actively replacing Splunk ' Guardium delivers 50'90% cost savings and is FedRAMP authorized. Merz is the branch head for both Sovereign Cloud and NS2 ' all IBM platform plays in this org start here.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "cedric bru": {

        "name": "Cedric Bru",

        "title": "CEO, Taulia",

        "org": "Taulia",

        "priority": "Low",

        "marker": None,

        "notes": "CEO, Taulia (SAP fintech subsidiary). Reports to Manoj Swaminathan. Taulia is SAP's embedded finance and supply chain finance platform. IBM watsonx.governance for financial AI compliance and watsonx Orchestrate for automated finance workflows are relevant as Taulia scales its AI capabilities. Currently shows zero direct reports in our data ' monitor for new contacts as the org is mapped.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "jonathan von rueden": {

        "name": "Jonathan Von Rueden",

        "title": "Development Senior Executive, SAP Business AI",

        "org": "Business AI",

        "priority": "High",

        "marker": "RED",

        "notes": "Development Senior Executive, SAP Business AI. Reports to Philipp Herzig. Team size: 1,053. Email: jonathan.von.rueden@sap.com. Location: Frankfurt. Per 8/17 and 8/24 notes: Von Rueden's org owns the SAP Business AI Platform spanning BDC and BTP ' this is where Joule is built and shipped. Andy Wei confirmed Joule is 3'4 years behind watsonx Orchestrate for A2A. Von Rueden reports directly to Herzig (CTO) and is the highest-value technical target for the Orchestrate competitive displacement play. Eliot Frederiksen is pursuing him for lunch in Palo Alto Sept 9. Dirk Basenach has been blocking Herzig access ' Von Rueden is the workaround path. High urgency.",

        "owner": "Ryan Sorber",

        "products": "watsonx Orchestrate / watsonx",

        "schedule": "N/A",

        "children": []

    },

    "andre wenz": {

        "name": "Andre Wenz",

        "title": "Head of SAP Signavio Product & Engineering",

        "org": "Signavio",

        "priority": "Low",

        "marker": None,

        "notes": "Head of SAP Signavio Product & Engineering. Reports to Philipp Herzig. Team size: 952. Signavio is SAP's business process intelligence and mining platform. IBM watsonx.governance is the natural AI governance layer for process mining outputs; watsonx Orchestrate can automate the downstream process improvements Signavio identifies. Depth-3 contact assigned to Eliot Frederiksen. Entry point for governance and process automation discussions at the Herzig org level.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "dominik rose": {

        "name": "Dominik Rose",

        "title": "Head of SAP LeanIX Product & Engineering",

        "org": "LeanIX",

        "priority": "Low",

        "marker": None,

        "notes": "Head of SAP LeanIX Product & Engineering. Reports to Philipp Herzig. Team size: 422. LeanIX is SAP's enterprise architecture management (EAM) platform. IBM watsonx.governance maps directly to the AI governance needs of enterprise architecture tooling. LeanIX's EA data is a natural input for IBM Turbonomic cloud cost optimization. Depth-3 contact assigned to Eliot Frederiksen. Strategic for governance and cloud cost plays within the Herzig CTO org.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "irfan khan": {

        "name": "Irfan Khan",

        "title": "Head of Business Data Cloud & Insights Management",

        "org": "Business Data Cloud",

        "priority": "Low",

        "marker": None,

        "notes": "Head of Business Data Cloud & Insights Management. Reports to Philipp Herzig. Team size: 4,299. Per 7/20 notes: BDC (Irfan Khan ' Prakash Nanduri ' Milinda Vitharana) is the master data team for all of SAP. IBM watsonx.data Premium with GPU acceleration delivered 83% cost/time reduction ' this resonated strongly with the BDC team. Confluent has an OEM partnership with SAP Datasphere (BDC product) ' Mark (Confluent rep) is problematic but the partnership creates an IBM entry point. Khan is the branch head for the highest-data-volume org in the Herzig tree. Critical for watsonx.data Premium expansion.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.data Premium / watsonx.governance / Confluent",

        "schedule": "N/A",

        "children": []

    },

    "anirban majumdar": {

        "name": "Anirban Majumdar",

        "title": "Development Group Executive, CTO Org",

        "org": "CTO Org / Business AI",

        "priority": "Low",

        "marker": None,

        "notes": "Development Group Executive, CTO Org. Reports to Philipp Herzig. 16 direct / 71 team. Part of the CTO core org under Herzig. IBM watsonx.governance and watsonx Orchestrate are the primary AI governance and automation plays for this engineering group. Assigned to Eliot Frederiksen as depth-3 contact within the Herzig CTO tree.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "dagmar schaffner": {

        "name": "Dagmar Schffner",

        "title": "Development Group Executive, CTO Org",

        "org": "CTO Org / Business AI",

        "priority": "Low",

        "marker": None,

        "notes": "Development Group Executive, CTO Org. Reports to Philipp Herzig. 12 direct / 24 team. Part of the CTO core org under Herzig. IBM watsonx.governance and watsonx Orchestrate are the primary plays. Per 8/10 notes: AI security is top of Herzig's mind (per Tanya Burley) ' Schffner as a CTO Dev Group Executive would be involved in the security and governance architecture decisions. Assigned to Eliot Frederiksen.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "eva klingbeil": {

        "name": "Eva Klingbeil",

        "title": "Development Group Executive, CTO Org",

        "org": "CTO Org / Business AI",

        "priority": "Low",

        "marker": None,

        "notes": "Development Group Executive, CTO Org. Reports to Philipp Herzig. Part of the CTO core org. Per 8/10 notes: Herzig's AI security focus (per Tanya Burley) creates opportunities for IBM AI governance across all CTO org engineering groups. Klingbeil is a development executive in this tree ' IBM watsonx.governance for AI security guardrails is the primary entry point. Assigned to Eliot Frederiksen.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "georg kniese": {

        "name": "Georg Kniese",

        "title": "Development Group Executive, CTO Org",

        "org": "CTO Org / Business AI",

        "priority": "Low",

        "marker": None,

        "notes": "Development Group Executive, CTO Org. Reports to Philipp Herzig. 4 direct / 55 team. Part of Herzig's CTO org with a relatively large team (55 total). IBM watsonx.governance and watsonx Orchestrate for AI automation are primary plays. Larger team size suggests platform or infrastructure scope ' Turbonomic for cloud cost optimization may also be relevant. Assigned to Eliot Frederiksen.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "gunther rothermel": {

        "name": "Gunther Rothermel",

        "title": "Development Group Executive, CTO Org",

        "org": "CTO Org / Business AI",

        "priority": "Low",

        "marker": None,

        "notes": "Development Group Executive, CTO Org. Reports to Philipp Herzig. 16 direct / 526 team. Per 8/10 notes: Herzig's AI security priority (per Tanya Burley) is relevant across all CTO engineering groups. Rothermel leads the largest team (526) among the CTO Dev Group Executives ' his org likely spans infrastructure and platform engineering where IBM Turbonomic, watsonx.governance, and IBM Bob are all relevant. Highest-priority CTO Org executive after Khan/Von Rueden by team size.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "kai muhlbauer": {

        "name": "Kai Mhlbauer",

        "title": "Development Group Executive, CTO Org",

        "org": "CTO Org / Business AI",

        "priority": "Low",

        "marker": None,

        "notes": "Development Group Executive, CTO Org. Reports to Philipp Herzig. Per 8/10 notes: AI security is top of Herzig's mind (per Tanya Burley, chief of staff). Mhlbauer is a CTO Dev Group Executive with direct line to Herzig ' IBM watsonx.governance for AI security guardrails and watsonx Orchestrate for AI workflow automation are the core plays. Assigned to Eliot Frederiksen.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "sophia mendelsohn": {

        "name": "Sophia Mendelsohn",

        "title": "Development Group Executive, CTO Org",

        "org": "CTO Org / Business AI",

        "priority": "Low",

        "marker": None,

        "notes": "Development Group Executive, CTO Org. Reports to Philipp Herzig. 8 direct / 105 team. Per 8/10 notes: AI security is Herzig's top priority (per Tanya Burley). Mendelsohn leads a 105-person engineering group ' at this scale, IBM watsonx.governance for AI compliance guardrails and IBM Bob for developer productivity acceleration are both directly relevant. Assigned to Eliot Frederiksen.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "tanja birli": {

        "name": "Tanja Birli",

        "title": "Development Group Executive, CTO Org",

        "org": "CTO Org / Business AI",

        "priority": "Low",

        "marker": None,

        "notes": "Development Group Executive, CTO Org. Reports to Philipp Herzig. 10 direct / 10 team. Part of the CTO core org. IBM watsonx.governance and watsonx Orchestrate are primary plays across all Herzig CTO engineering groups. Per 8/10 notes: AI security focus (per Tanya Burley) creates governance entry points. Assigned to Eliot Frederiksen.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "yaad oren": {

        "name": "Yaad Oren",

        "title": "Development Group Executive, CTO Org",

        "org": "CTO Org / Business AI",

        "priority": "Low",

        "marker": None,

        "notes": "Development Group Executive, CTO Org. Reports to Philipp Herzig. 17 direct / 211 team. Leads a large engineering org (211 total) under Herzig. Per 8/10 notes: AI security is Herzig's top concern (per Tanya Burley). Oren's team size suggests broad platform scope ' IBM watsonx.governance for AI governance at scale, Turbonomic for cloud cost optimization, and IBM Bob for developer productivity are all relevant. Second-largest CTO Org team after Rothermel. Assigned to Eliot Frederiksen.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "customer & cloud ops (cdx)": {

        "name": "Customer & Cloud Ops (CDX)",

        "title": "Customer & Cloud Ops Infrastructure Branches (Unchanged)",

        "org": "Customer & Cloud Ops",

        "priority": "Low",

        "marker": None,

        "notes": "Customer & Cloud Ops (CDX) ' infrastructure branches reporting to Thomas Saueressig. Per 7/20 notes: AIOps POC budget was approved in CDX but is stuck on leadership mobilization. Mona Mohan George leads CDX; Deo Caruana is the SRE decision-maker. IBM Turbonomic, Guardium Compliance Auto, and Confluent are all active fits for this org. CDX represents the largest single-org IBM opportunity in the near term ' unblocking the AIOps POC is the immediate action item.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    }

}

# 8 SAP Executive Board Members reporting to Christian Klein

board_members = {

    "manoj swaminathan": {

        "name": "Manoj Swaminathan",

        "title": "President & CPO, SAP Autonomous Suite",

        "org": "Autonomous Suite",

        "priority": "Low",

        "marker": None,

        "notes": "President & CPO, SAP Autonomous Suite. Reports directly to Christian Klein, CEO. As board-level CPO, Swaminathan sets the product strategy for SAP's autonomous and AI-powered suite ' IBM watsonx Orchestrate (for autonomous workflow) and watsonx.governance (for responsible AI) are aligned to the Autonomous Suite vision. All branch heads (Lawrence Martin, Siva Sundaresan, Vijay Seethapathy, Cedric Bru, Martin Merz) report through him. Eliot Frederiksen manages top-level relationship.",

        "owner": "Jerome Carlson (ATL)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "philipp herzig": {

        "name": "Philipp Herzig",

        "title": "SVP, CTO & Chief AI Officer, SAP",

        "org": "CTO Org / Business AI",

        "priority": "Low",

        "marker": None,

        "notes": "SVP, CTO & Chief AI Officer. Direct report to Christian Klein, CEO. Per 8/10 notes: AI security is Herzig's top concern (per Tanya Burley, his chief of staff). Per 8/17 notes: Eliot Frederiksen is pursuing Herzig for lunch in Palo Alto Sept 9. Dirk Basenach is a gatekeeper who has been blocking Herzig access ' Von Rueden is the workaround. Herzig owns the entire IBM-SAP technical strategy conversation, including the ESA (~$40M restatement risk) and all Business AI, BDC, Signavio, LeanIX, and Sovereign Cloud plays. Rob Thomas (IBM CCO) is the executive counterpart. Highest-priority SAP executive after Christian Klein.",

        "owner": "Rob Thomas (IBM CCO)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "thomas saueressig": {

        "name": "Thomas Saueressig",

        "title": "Chief Customer Officer, Executive Board, SAP",

        "org": "Customer & Cloud Ops",

        "priority": "Low",

        "marker": None,

        "notes": "Chief Customer Officer, Executive Board. Leading Customer & Cloud Ops (~60% of org) and Run consolidation. Per 7/20 notes: CDX (under Vijay Seethapathy, reporting to Saueressig) has an approved AIOps POC budget that is stuck ' IBM needs to mobilize through Saueressig to unlock it. Turbonomic for cloud cost consolidation and watsonx Orchestrate for customer operations automation are the primary plays at this level. Rob Thomas (IBM CCO) is the executive counterpart for board-level relationship.",

        "owner": "Rob Thomas (IBM CCO)",

        "products": "watsonx.governance / watsonx Orchestrate / Turbonomic",

        "schedule": "N/A",

        "children": []

    },

    "sebastian steinhaeuser": {

        "name": "Sebastian Steinhaeuser",

        "title": "Chief Strategy Officer & COO, Executive Board, SAP",

        "org": "Strategy & Operations",

        "priority": "Low",

        "marker": None,

        "notes": "Chief Strategy Officer & COO, Executive Board. Leading Corporate Strategy Execution, CIO Office, and organizational simplification. Per 8/24 notes: ESA audit issue creates ~$40M restatement risk for SAP from undeployed software ' Steinhaeuser as COO owns the corporate execution track where IBM can help SAP accelerate internal deployments. CIO Office (Benjamin Blau, Christian Freytag) reports through him. IBM's SAP-runs-SAP play (IBM helping SAP use its own software faster) is the strategic frame here. Joanne Wright (IBM CIO) is the executive counterpart.",

        "owner": "Joanne Wright (IBM CIO)",

        "products": "watsonx.governance / watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "dominik asam": {

        "name": "Dominik Asam",

        "title": "Chief Financial Officer, Executive Board, SAP",

        "org": "Finance",

        "priority": "Low",

        "marker": None,

        "notes": "Chief Financial Officer, Executive Board. Leading global Finance, Investor Relations, and corporate cost structure. Per 8/24 notes: ESA (~$40M restatement risk) is a CFO-level issue ' Asam is ultimately accountable for revenue recognition and would feel urgency around any undeployed software creating audit exposure. IBM Turbonomic for cloud FinOps and watsonx.governance for financial AI compliance are the plays at this level. James Kavanaugh (IBM CFO) is the executive counterpart ' C-to-C financial peer conversation.",

        "owner": "James Kavanaugh (IBM CFO)",

        "products": "Turbonomic / watsonx.governance",

        "schedule": "N/A",

        "children": []

    },

    "gina vargiu-breuer": {

        "name": "Gina Vargiu-Breuer",

        "title": "Chief CHRO / Chief People Officer, Executive Board, SAP",

        "org": "HR / People Org",

        "priority": "Low",

        "marker": None,

        "notes": "Chief HR Officer & Chief People Officer, Executive Board. Leading global HR and people organization. IBM watsonx Orchestrate is the leading AI platform for HR workflow automation ' directly competes with SAP Joule (which Andy Wei confirmed is 3'4 years behind Orchestrate for A2A). SuccessFactors is SAP's HCM product that Vargiu-Breuer would leverage internally. Nickle LaMoreaux (IBM CHRO) is the executive counterpart ' C-to-C HR peer conversation creates a natural opening.",

        "owner": "Nickle LaMoreaux (IBM CHRO)",

        "products": "watsonx Orchestrate",

        "schedule": "N/A",

        "children": []

    },

    "ada agrait": {

        "name": "Ada Agrait",

        "title": "Chief Marketing Officer, Executive Board, SAP",

        "org": "Marketing",

        "priority": "Low",

        "marker": None,

        "notes": "Chief Marketing Officer, Executive Board. Leading global Marketing and Sapphire events including SAP Sapphire. IBM watsonx.governance for AI-generated content compliance is directly relevant as SAP scales AI-assisted marketing. IBM is a prominent Sapphire sponsor ' Agrait controls the venue and partnership relationships. Scott Baker (IBM CMO) is the executive counterpart. Marketing-to-marketing peer relationship creates natural co-branding and co-marketing opportunities around the IBM-SAP partnership narrative.",

        "owner": "Scott Baker (IBM CMO)",

        "products": "watsonx.governance",

        "schedule": "N/A",

        "children": []

    }

}

# Christian Klein (Absolute Root CEO)

root_node = {

    "name": "Christian Klein",

    "title": "CEO & Chair of the Executive Board, SAP SE",

    "org": "Executive Board",

    "priority": "Low",

    "marker": None,

    "notes": "CEO & Chair of the Executive Board, SAP SE. Arvind Krishna (IBM CEO) is the executive counterpart. Per 8/24 notes: ESA (~$40M restatement risk from undeployed IBM software) is an Arvind-Christian conversation ' the ESA creates deployment urgency that IBM should leverage across all LOBs. Christian Klein's strategic focus on Business AI (Joule, BTP) and cloud consolidation aligns with IBM's entire portfolio. This is the apex relationship ' all board-member and branch-head plays roll up to the Klein-Krishna executive alignment.",

    "owner": "Arvind Krishna / Jerome Carlson",

    "products": "Enterprise Solutions Agreement (ESA)",

    "schedule": "N/A",

    "children": []

}

# Create JS nodes

nodes = {}

for name_lower, c in contacts.items():

    nodes[name_lower] = {

        "name": c["name"],

        "title": c["title"],

        "org": c["org"],

        "priority": c["priority"],

        "marker": c["marker"],

        "notes": c["notes"],

        "owner": c["owner"],

        "products": c["products"],

        "schedule": c["schedule"],

        "branch": c["branch"],

        "reports_to": c["reports_to"],

        "verification": c.get("verification") or "Not yet run through the public-verification pass.",

        "verification_note": c.get("verification_note") or "No verification notes available.",

        "children": []

    }

# Link nodes

for name_lower, node in nodes.items():

    mgr = node["reports_to"]

    if not mgr:

        org_text = (node.get("org") or "").lower()

        branch_text = (node.get("branch") or "").lower()

        if "cdx" in org_text or "sre" in org_text or "cloud ops" in org_text or "pmo" in org_text:

            branch_heads["customer & cloud ops (cdx)"]["children"].append(node)

        elif "ai" in org_text or "cto" in org_text or "product management" in org_text:

            board_members["philipp herzig"]["children"].append(node)

        else:

            if node["branch"] == "(unchanged branch, not part of the 6-anchor rebuild)" or not node["branch"]:

                board_members["philipp herzig"]["children"].append(node)

            else:

                b_lower = node["branch"].lower()

                if b_lower in branch_heads:

                    branch_heads[b_lower]["children"].append(node)

                else:

                    board_members["philipp herzig"]["children"].append(node)

    else:

        mgr_lower = mgr.lower()

        if mgr_lower in nodes:

            nodes[mgr_lower]["children"].append(node)

        elif mgr_lower in board_members:

            board_members[mgr_lower]["children"].append(node)

        elif mgr_lower in branch_heads:

            branch_heads[mgr_lower]["children"].append(node)

        else:

            matched = False

            for bh_lower in branch_heads:

                if bh_lower in mgr_lower:

                    branch_heads[bh_lower]["children"].append(node)

                    matched = True

                    break

            if not matched:

                org_text = (node.get("org") or "").lower()

                if "cdx" in org_text or "sre" in org_text or "cloud ops" in org_text or "pmo" in org_text:

                    branch_heads["customer & cloud ops (cdx)"]["children"].append(node)

                else:

                    board_members["philipp herzig"]["children"].append(node)

# Add board members under Christian Klein

for bm_name, bm_node in board_members.items():

    root_node["children"].append(bm_node)

# Attach branch heads to correct board members

for bh_name, bh_node in branch_heads.items():

    if bh_name in ("jonathan von rueden", "andre wenz", "dominik rose", "irfan khan",

                   "anirban majumdar", "dagmar schaffner", "eva klingbeil", "georg kniese",

                   "gunther rothermel", "kai muhlbauer", "sophia mendelsohn", "tanja birli", "yaad oren"):

        board_members["philipp herzig"]["children"].append(bh_node)

    elif bh_name == "customer & cloud ops (cdx)":

        board_members["thomas saueressig"]["children"].append(bh_node)

    else:

        board_members["manoj swaminathan"]["children"].append(bh_node)

# Recursively sort children under each node by LOB (org), then alphabetically by name!

# This ensures people in the same org are always grouped together side-by-side

def sort_tree_children(node):

    node["children"].sort(key=lambda x: (x.get("org") or "", x.get("name") or ""))

    for child in node["children"]:

        sort_tree_children(child)

sort_tree_children(root_node)

# "" Assign Eliot Frederiksen to all contacts at Fang Chang's depth (3) and above ""

# Fang Chang sits at depth 3 (CEO=0, Board=1, Branch heads=2, Fang's peers=3)

def _assign_eliot_above_fang(node, depth=0, max_depth=3):

    if depth <= max_depth:

        node["owner"] = "Eliot Frederiksen (IBM Client Executive)"

    for child in node.get("children", []):

        _assign_eliot_above_fang(child, depth + 1, max_depth)

_assign_eliot_above_fang(root_node)

# "" Assign owners for Vijay Seethapathy's org """"""""""""""""""""""""""""""""""

# Rules:

#   Vijay himself              ' already set to Eliot (depth 3) " leave as-is

#   Vijay's direct reports     ' Jerome Carlson (ATL)

#   Mona's level (depth 5) and their direct reports (depth 6) ' Jerome Carlson (ATL)

#   Everything below depth 6   ' Josh McClure

#

# We walk the tree starting from the vijay node.

def _find_node(node, name_lower):

    if node["name"].lower() == name_lower:

        return node

    for child in node.get("children", []):

        result = _find_node(child, name_lower)

        if result:

            return result

    return None

def _assign_vijay_org_owners(node, vijay_depth, current_depth):

    """

    vijay_depth  = depth of Vijay in the full tree.

    current_depth = depth of `node`.

    relative depth = current_depth - vijay_depth:

      0 ' Vijay himself (skip, keep Eliot)

      1 ' Vijay's direct reports ' Jerome

      2 ' Mona's level ' Jerome

      3 ' Mona's direct reports ' Jerome

      4+ ' Josh McClure

    """

    rel = current_depth - vijay_depth

    if rel == 1:

        node["owner"] = "Jerome Carlson (ATL)"

    elif rel in (2, 3):

        node["owner"] = "Jerome Carlson (ATL)"

    elif rel >= 4:

        node["owner"] = "Josh McClure"

    for child in node.get("children", []):

        _assign_vijay_org_owners(child, vijay_depth, current_depth + 1)

def _find_depth(node, name_lower, current_depth=0):

    if node["name"].lower() == name_lower:

        return current_depth

    for child in node.get("children", []):

        result = _find_depth(child, name_lower, current_depth + 1)

        if result is not None:

            return result

    return None

vijay_node = _find_node(root_node, "vijay seethapathy")

if vijay_node:

    vijay_depth_in_tree = _find_depth(root_node, "vijay seethapathy") or 2

    _assign_vijay_org_owners(vijay_node, vijay_depth_in_tree, vijay_depth_in_tree)

# "" Assign owners/products for Lawrence Martin's org below Jerome's level """"""

# Jerome Carlson (ATL) is the ATL, so depth 3 = branch heads (Lawrence Martin's level).

# "Below Jerome's level" = depth > 3, i.e. everyone who reports to Lawrence Martin

# or deeper.  Assign Ryan Sorber + watsonx / IBM Bob / Manta.

LAWRENCE_PRODUCTS = "watsonx / IBM Bob (Code Assistant) / Manta (Data Lineage)"

def _assign_lawrence_org(node, lawrence_depth, current_depth):

    """rel depth 1+ (below Lawrence Martin himself) ' Ryan Sorber + watsonx/Bob/Manta products"""

    rel = current_depth - lawrence_depth

    if rel >= 1:

        node["owner"] = "Ryan Sorber"

        node["products"] = LAWRENCE_PRODUCTS

    for child in node.get("children", []):

        _assign_lawrence_org(child, lawrence_depth, current_depth + 1)

lawrence_node = _find_node(root_node, "lawrence martin")

if lawrence_node:

    lawrence_depth_in_tree = _find_depth(root_node, "lawrence martin") or 2

    _assign_lawrence_org(lawrence_node, lawrence_depth_in_tree, lawrence_depth_in_tree)

# -- Assign schedule weeks by priority --

# Relationship  ' N/A

# High          ' Week 1"4   (cycle through 1,2,3,4 as contacts are encountered)

# Medium        ' Week 5"12  (cycle through 512)

# Low           ' Week 12"30 (cycle through 1230)

# Owner default ' Jerome Carlson (ATL) if blank

import itertools

_high_weeks   = itertools.cycle(range(1, 5))       # 1-4

_medium_weeks = itertools.cycle(range(5, 13))      # 5-12

_low_weeks    = itertools.cycle(range(12, 31))     # 12-30

def _assign_schedule(node):

    p = node.get("priority", "Low")

    # owner default

    if not node.get("owner") or node["owner"].strip().lower() in ("", "unassigned"):

        node["owner"] = "Jerome Carlson (ATL)"

    if p == "Relationship":

        node["schedule"] = "N/A"

    elif p == "High":

        node["schedule"] = f"Week {next(_high_weeks)}"

    elif p == "Medium":

        node["schedule"] = f"Week {next(_medium_weeks)}"

    else:

        node["schedule"] = f"Week {next(_low_weeks)}"

    for child in node.get("children", []):

        _assign_schedule(child)

# Apply to the full tree (covers all nodes including branch heads & board members)

_assign_schedule(root_node)

# -- Email-derived owner overrides --

# Applied AFTER _assign_schedule so these take precedence over all rule-based

# defaults.  Each entry maps a contact name (lowercase) to the IBM seller who

# is the confirmed primary contact, determined by email thread analysis using

# the seniority ladder: Eliot > Jerome > Tanja > Ryan > Josh > Dirk > Werner

# > Yusuke > Candace > Kevin > anyone else.

_EMAIL_OWNER_OVERRIDES = {

    # Eliot is primary (direct thread originator or most senior IBM on thread)

    "chuck firkin":               "Eliot Frederiksen (IBM Client Executive)",

    "atul deo":                   "Eliot Frederiksen (IBM Client Executive)",

    "marc-oliver klein":          "Eliot Frederiksen (IBM Client Executive)",

    "ajit kaicker":               "Eliot Frederiksen (IBM Client Executive)",

    "gokul naidu":                "Eliot Frederiksen (IBM Client Executive)",

    "daniel beck":                "Eliot Frederiksen (IBM Client Executive)",

    "nick totten":                "Eliot Frederiksen (IBM Client Executive)",

    "balaji balasubramanian":     "Eliot Frederiksen (IBM Client Executive)",

    "collin clark":               "Eliot Frederiksen (IBM Client Executive)",

    # Jerome is primary (most senior IBM on thread, Josh or lower are leads)

    "santosh tej":                "Jerome Carlson (ATL)",

    "kumar sambhav":              "Jerome Carlson (ATL)",

    "konstantinos michalitsis":   "Jerome Carlson (ATL)",

    "anthony sanchez":            "Jerome Carlson (ATL)",

    "harsha ravuri":              "Jerome Carlson (ATL)",

    "gana sadasivam":             "Jerome Carlson (ATL)",

    "daniel cooley":              "Jerome Carlson (ATL)",

    "mehmet yurdal":              "Jerome Carlson (ATL)",

    "gregor tielsch":             "Jerome Carlson (ATL)",

    # Josh is primary (Josh led thread, Jerome/Eliot on CC only)

    "mona mohan george":          "Josh McClure",

    "john mazzolini":             "Josh McClure",

    "carey main":                 "Josh McClure",

    "jim rotan":                  "Josh McClure",

    "john jiang":                 "Josh McClure",

    "james reidy":                "Josh McClure",

    "jackson borges":             "Josh McClure",

    "mark devita":                "Josh McClure",

    "adam levinson":              "Josh McClure",

    "ryan duncan":                "Josh McClure",

    # Ryan is primary (Ryan was sole IBM sender on thread)

    "jonathan von rueden":        "Ryan Sorber",

    "brian morrison":             "Ryan Sorber",

    "samantha combs":             "Ryan Sorber",

    # Dirk is primary (Dirk sent BDC summary; Eliot on CC ' Eliot by seniority

    # but Dirk is the technical lead relationship owner " apply Dirk)

    "marc geall":                 "Dirk Basenach",

    "h nair":                     "Dirk Basenach",

}

def _apply_email_owner_overrides(node):

    key = node["name"].lower()

    if key in _EMAIL_OWNER_OVERRIDES:

        node["owner"] = _EMAIL_OWNER_OVERRIDES[key]

    for child in node.get("children", []):

        _apply_email_owner_overrides(child)

_apply_email_owner_overrides(root_node)

# Also apply overrides to contacts dict directly (for extra_contacts not in tree)

for key, owner in _EMAIL_OWNER_OVERRIDES.items():

    if key in contacts:

        contacts[key]["owner"] = owner

# Sync back into flat contacts dict so call_list gets updated values too

def _sync_contacts_from_tree(node):

    key = node["name"].lower()

    if key in contacts:

        contacts[key]["schedule"] = node["schedule"]

        contacts[key]["owner"] = node["owner"]

    for child in node.get("children", []):

        _sync_contacts_from_tree(child)

_sync_contacts_from_tree(root_node)

# Create Call List data sorted in order of priority:

# Relationship (Amber/Orange) -> High (Red) -> Medium (Yellow) -> Low (Grey)

priority_order = {"Relationship": 1, "High": 2, "Medium": 3, "Low": 4}

call_list = []

for k, c in contacts.items():

    call_list.append({

        "name": c["name"],

        "title": c["title"],

        "org": c["org"],

        "priority": c["priority"],

        "notes": c["notes"],

        "owner": c["owner"],

        "products": c["products"],

        "schedule": c["schedule"],

        "branch": c["branch"],

        "verification": c.get("verification") or "Not yet run through the public-verification pass.",

        "verification_note": c.get("verification_note") or "No verification notes available."

    })

call_list.sort(key=lambda x: priority_order.get(x["priority"], 9))

# Write out the full html page!

html_content = f"""<!DOCTYPE html>

<html lang="en">

<head>

    <meta charset="UTF-8">

    <meta name="viewport" content="width=device-width, initial-scale=1.0">

    <title>SAP US Account Priority Technical Outreach & Org Chart</title>

    <style>

        :root {{

            --bg-color: #ffffff;

            --surface-color: #f7f8fa;

            --border-color: #cbd5e1;

            --text-color: #1f2328;

            --muted-color: #57606a;

            --connector-color: #475569;

            /* Priority Colors */

            --color-relationship-bg: #f0fdf4;

            --color-relationship-border: #16a34a;

            --color-relationship-text: #14532d;

            --color-high-bg: #fef2f2;

            --color-high-border: #ef4444;

            --color-high-text: #b91c1c;

            --color-medium-bg: #fefce8;

            --color-medium-border: #facc15;

            --color-medium-text: #a16207;

            --color-low-bg: #f3f4f6;

            --color-low-border: #9ca3af;

            --color-low-text: #374151;

            --font-family: -apple-system, "Segoe UI", system-ui, sans-serif;

        }}

        * {{

            box-sizing: border-box;

            margin: 0;

            padding: 0;

        }}

        body {{

            font-family: var(--font-family);

            background-color: var(--bg-color);

            color: var(--text-color);

            line-height: 1.5;

            padding: 24px;

        }}

        .container {{

            max-width: 1400px;

            margin: 0 auto;

        }}

        header {{

            margin-bottom: 24px;

            border-bottom: 2px solid var(--border-color);

            padding-bottom: 16px;

        }}

        h1 {{

            font-size: 24px;

            font-weight: 700;

            margin-bottom: 4px;

            color: #1a365d;

        }}

        .subtitle {{

            color: var(--muted-color);

            font-size: 14px;

        }}

        /* Tabs Navigation */

        .tabs {{

            display: flex;

            gap: 8px;

            margin-bottom: 24px;

            border-bottom: 1px solid var(--border-color);

            padding-bottom: 1px;

        }}

        .tab-btn {{

            background: none;

            border: none;

            padding: 10px 16px;

            font-size: 15px;

            font-weight: 600;

            color: var(--muted-color);

            cursor: pointer;

            border-radius: 6px 6px 0 0;

            border: 1px solid transparent;

            margin-bottom: -1px;

            transition: all 0.2s ease;

        }}

        .tab-btn:hover {{

            color: var(--text-color);

            background-color: var(--surface-color);

        }}

        .tab-btn.active {{

            color: #1a365d;

            background-color: var(--bg-color);

            border-color: var(--border-color) var(--border-color) transparent var(--border-color);

            border-bottom: 2px solid #1a365d;

        }}

        .tab-content {{

            display: none;

        }}

        .tab-content.active {{

            display: block;

        }}

        /* Controls / Search Bar */

        .controls-row {{

            display: flex;

            justify-content: space-between;

            align-items: center;

            gap: 16px;

            margin-bottom: 20px;

            flex-wrap: wrap;

        }}

        .search-container {{

            position: relative;

            flex-grow: 1;

            max-width: 400px;

        }}

        .search-input {{

            width: 100%;

            padding: 10px 12px 10px 36px;

            border: 1px solid var(--border-color);

            border-radius: 6px;

            font-size: 14px;

            background-color: var(--surface-color);

            outline: none;

            transition: all 0.2s ease;

        }}

        .search-input:focus {{

            border-color: #3b82f6;

            background-color: #fff;

            box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.15);

        }}

        .search-icon {{

            position: absolute;

            left: 12px;

            top: 50%;

            transform: translateY(-50%);

            color: var(--muted-color);

            pointer-events: none;

        }}

        .btn-group {{

            display: flex;

            gap: 8px;

        }}

        .btn {{

            background-color: var(--surface-color);

            border: 1px solid var(--border-color);

            padding: 8px 14px;

            font-size: 14px;

            font-weight: 500;

            border-radius: 6px;

            cursor: pointer;

            transition: all 0.1s ease;

            display: inline-flex;

            align-items: center;

            gap: 6px;

        }}

        .btn:hover {{

            background-color: #f1f2f4;

            border-color: #d1d5db;

        }}

        .btn.primary {{

            background-color: #1a365d;

            color: #fff;

            border-color: #1a365d;

        }}

        .btn.primary:hover {{

            background-color: #142a4a;

        }}

        /* Org Chart Styling with Pan & Zoom Viewport */

        .org-tree-wrapper {{

            overflow: hidden; /* managed via custom click & drag pan and scroll wheel zoom */

            border: 1px solid var(--border-color);

            border-radius: 8px;

            background-color: var(--surface-color);

            padding: 40px;

            height: 650px;

            position: relative;

            cursor: grab;

            user-select: none;

        }}

        .org-tree-wrapper:active {{

            cursor: grabbing;

        }}

        /* Horizontal Tree Root Container */

        .org-tree-root {{

            display: inline-block;

            width: max-content;

            transform-origin: top center;

            transition: transform 0.05s ease-out;

            position: absolute;

            top: 40px;

            left: 50%;

            transform: translateX(-50%) scale(0.85); /* start centered and slightly zoomed out */

        }}

        .tree-children {{

            display: flex;

            flex-direction: row;

            flex-wrap: wrap;

            justify-content: center;

            align-items: flex-start;

            position: relative;

            padding-top: 24px;

            gap: 16px;

            max-width: 1780px;

            margin: 0 auto;

            transition: all 0.3s ease;

        }}

        /* Vertical parent line entering the children container */

        .tree-children::before {{

            content: '';

            position: absolute;

            top: 0;

            left: 50%;

            border-left: 2px solid var(--connector-color);

            width: 0;

            height: 24px;

        }}

        .tree-node-container {{

            display: flex;

            flex-direction: column;

            align-items: center;

            position: relative;

        }}

        /* Sibling connector line entering each child node or LOB group box */

        .tree-children > .tree-node-container::before,

        .tree-children > .tree-node-container::after,

        .tree-children > .lob-group-box::before,

        .tree-children > .lob-group-box::after {{

            content: '';

            position: absolute;

            top: 0;

            width: 50%;

            border-top: 2px solid var(--connector-color);

        }}

        .tree-children > .tree-node-container::before,

        .tree-children > .lob-group-box::before {{

            left: 0;

        }}

        .tree-children > .tree-node-container::after,

        .tree-children > .lob-group-box::after {{

            right: 0;

        }}

        /* Sibling line edges */

        .tree-children > .tree-node-container:first-child::before,

        .tree-children > .lob-group-box:first-child::before {{

            border: none;

        }}

        .tree-children > .tree-node-container:last-child::after,

        .tree-children > .lob-group-box:last-child::after {{

            border: none;

        }}

        .tree-children > .tree-node-container:only-child::before,

        .tree-children > .tree-node-container:only-child::after,

        .tree-children > .lob-group-box:only-child::before,

        .tree-children > .lob-group-box:only-child::after {{

            border: none;

        }}

        /* Internal connector lines for children inside an LOB group box */

        .lob-group-box > .tree-node-container::before,

        .lob-group-box > .tree-node-container::after {{

            content: '';

            position: absolute;

            top: -24px;

            width: 50%;

            border-top: 2px solid var(--connector-color);

        }}

        .lob-group-box > .tree-node-container::before {{

            left: 0;

        }}

        .lob-group-box > .tree-node-container::after {{

            right: 0;

        }}

        .lob-group-box > .tree-node-container:first-child::before {{

            border: none;

        }}

        .lob-group-box > .tree-node-container:last-child::after {{

            border: none;

        }}

        .lob-group-box > .tree-node-container:only-child::before,

        .lob-group-box > .tree-node-container:only-child::after {{

            border: none;

        }}

        /* Vertical connector stub from card down to children row " hidden when collapsed */

        .tree-node-card-connector {{

            display: none;

        }}

        /* LOB Group Dotted Boundary Box Styling */

        .lob-group-box {{

            display: flex;

            flex-direction: row;

            flex-wrap: wrap;

            justify-content: center;

            align-items: flex-start;

            border: 2px dashed #9ca3af;

            border-radius: 12px;

            padding: 28px 16px 16px 16px;

            margin: 0 12px;

            background-color: rgba(243, 244, 246, 0.4);

            position: relative;

            gap: 16px;

            max-width: 1780px;

        }}

        .lob-group-box-title {{

            position: absolute;

            top: -10px;

            left: 16px;

            font-size: 10px;

            font-weight: 800;

            padding: 2px 8px;

            background-color: #ffffff;

            border-radius: 4px;

            text-transform: uppercase;

            letter-spacing: 0.5px;

            border: 1px dashed #9ca3af;

            white-space: nowrap;

        }}

        /* Custom color styles for each primary LOB group */

        .lob-group-box.lob-Ariba {{ border-color: #3b82f6; background-color: rgba(59, 130, 246, 0.03); }}

        .lob-group-box.lob-Ariba .lob-group-box-title {{ border-color: #3b82f6; color: #1d4ed8; }}

        .lob-group-box.lob-NS2 {{ border-color: #10b981; background-color: rgba(16, 185, 129, 0.03); }}

        .lob-group-box.lob-NS2 .lob-group-box-title {{ border-color: #10b981; color: #047857; }}

        .lob-group-box.lob-SuccessFactors {{ border-color: #8b5cf6; background-color: rgba(139, 92, 246, 0.03); }}

        .lob-group-box.lob-SuccessFactors .lob-group-box-title {{ border-color: #8b5cf6; color: #6d28d9; }}

        .lob-group-box.lob-Concur {{ border-color: #14b8a6; background-color: rgba(20, 184, 166, 0.03); }}

        .lob-group-box.lob-Concur .lob-group-box-title {{ border-color: #14b8a6; color: #0f766e; }}

        .lob-group-box.lob-CX {{ border-color: #f97316; background-color: rgba(249, 115, 22, 0.03); }}

        .lob-group-box.lob-CX .lob-group-box-title {{ border-color: #f97316; color: #c2410c; }}

        .lob-group-box.lob-CDX {{ border-color: #ec4899; background-color: rgba(236, 72, 153, 0.03); }}

        .lob-group-box.lob-CDX .lob-group-box-title {{ border-color: #ec4899; color: #be185d; }}

        .lob-group-box.lob-SovereignCloud {{ border-color: #ef4444; background-color: rgba(239, 68, 68, 0.03); }}

        .lob-group-box.lob-SovereignCloud .lob-group-box-title {{ border-color: #ef4444; color: #b91c1c; }}

        .lob-group-box.lob-BusinessNetwork {{ border-color: #6366f1; background-color: rgba(99, 102, 241, 0.03); }}

        .lob-group-box.lob-BusinessNetwork .lob-group-box-title {{ border-color: #6366f1; color: #4338ca; }}

        /* Override to fallback to single nowrap horizontal row if any child is expanded */

        .tree-children.has-expanded-child,

        .lob-group-box.has-expanded-child {{

            flex-wrap: nowrap !important;

            max-width: none !important;

            width: max-content !important;

        }}

        /* Floating Zoom & Pan Controls styling */

        .zoom-controls {{

            position: absolute;

            bottom: 24px;

            right: 24px;

            display: flex;

            flex-direction: column;

            gap: 8px;

            background-color: var(--bg-color);

            border: 1px solid var(--border-color);

            border-radius: 8px;

            padding: 8px;

            box-shadow: 0 4px 12px rgba(0,0,0,0.1);

            z-index: 10;

        }}

        .zoom-btn {{

            width: 36px;

            height: 36px;

            border-radius: 6px;

            border: 1px solid var(--border-color);

            background-color: var(--surface-color);

            font-size: 18px;

            font-weight: 700;

            cursor: pointer;

            display: flex;

            align-items: center;

            justify-content: center;

            transition: all 0.1s ease;

        }}

        .zoom-btn:hover {{

            background-color: #f1f2f4;

            border-color: #d1d5db;

        }}

        .zoom-info {{

            font-size: 10px;

            font-weight: 700;

            color: var(--muted-color);

            text-align: center;

            margin-top: 4px;

        }}

        /* CARD STYLING */

        .node-card {{

            width: 240px;

            background-color: var(--bg-color);

            border: 2px solid var(--border-color);

            border-radius: 8px;

            padding: 12px;

            position: relative;

            box-shadow: 0 1px 3px rgba(0,0,0,0.05);

            transition: all 0.2s ease;

            text-align: center;

            cursor: pointer;

            user-select: none;

        }}

        .node-card:hover {{

            box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1), 0 2px 4px -1px rgba(0,0,0,0.06);

            transform: translateY(-2px);

        }}

        .node-card.highlighted {{

            border-color: #2563eb !important;

            box-shadow: 0 0 0 4px rgba(37, 99, 235, 0.25) !important;

            transform: scale(1.05);

            z-index: 10;

        }}

        /* Priority card variations */

        .node-card.priority-Relationship {{

            background-color: var(--color-relationship-bg);

            border-color: var(--color-relationship-border);

        }}

        .node-card.priority-High {{

            background-color: var(--color-high-bg);

            border-color: var(--color-high-border);

        }}

        .node-card.priority-Medium {{

            background-color: var(--color-medium-bg);

            border-color: var(--color-medium-border);

        }}

        .node-card.priority-Low {{

            background-color: var(--color-low-bg);

            border-color: var(--color-low-border);

        }}

        /* Inside card elements */

        .card-priority-pill {{

            position: absolute;

            top: -10px;

            left: 50%;

            transform: translateX(-50%);

            font-size: 10px;

            font-weight: 700;

            padding: 2px 8px;

            border-radius: 12px;

            text-transform: uppercase;

            letter-spacing: 0.5px;

            border: 1px solid transparent;

        }}

        .priority-Relationship .card-priority-pill {{

            background-color: var(--color-relationship-border);

            color: #ffffff;

        }}

        .priority-High .card-priority-pill {{

            background-color: var(--color-high-border);

            color: #ffffff;

        }}

        .priority-Medium .card-priority-pill {{

            background-color: var(--color-medium-border);

            color: var(--color-medium-text);

        }}

        .priority-Low .card-priority-pill {{

            background-color: var(--color-low-border);

            color: #ffffff;

        }}

        .card-name {{

            font-size: 14px;

            font-weight: 700;

            color: var(--text-color);

            margin-top: 4px;

            margin-bottom: 2px;

        }}

        .card-title {{

            font-size: 11px;

            color: var(--muted-color);

            font-weight: 600;

            line-height: 1.3;

            margin-bottom: 4px;

            min-height: 28px;

            display: flex;

            align-items: center;

            justify-content: center;

        }}

        .card-org {{

            font-size: 10px;

            font-weight: 700;

            color: #475569;

            text-transform: uppercase;

            letter-spacing: 0.3px;

        }}

        /* Inline Card Name Toggle Button */

        .card-name-toggle {{

            display: inline-flex;

            align-items: center;

            justify-content: center;

            width: 18px;

            height: 18px;

            border-radius: 4px;

            border: 1px solid var(--border-color);

            background-color: var(--surface-color);

            font-size: 11px;

            font-weight: 800;

            cursor: pointer;

            margin-left: 8px;

            vertical-align: middle;

            color: var(--muted-color);

            line-height: 1;

            transition: all 0.1s ease;

        }}

        .card-name-toggle:hover {{

            background-color: #f1f2f4;

            border-color: #d1d5db;

            color: var(--text-color);

            transform: scale(1.1);

        }}

        /* Table Row Toggle Styles */

        .table-row-toggle-btn {{

            display: inline-flex;

            align-items: center;

            justify-content: center;

            width: 18px;

            height: 18px;

            border-radius: 4px;

            border: 1px solid var(--border-color);

            background-color: var(--surface-color);

            font-size: 11px;

            font-weight: 800;

            cursor: pointer;

            margin-right: 8px;

            vertical-align: middle;

            color: var(--muted-color);

            line-height: 1;

            transition: all 0.1s ease;

        }}

        .table-row-toggle-btn:hover {{

            background-color: #f1f2f4;

            border-color: #d1d5db;

            color: var(--text-color);

        }}

        /* Edit Form Controls */

        .form-input {{

            width: 100%;

            padding: 8px 10px;

            border: 1px solid var(--border-color);

            border-radius: 6px;

            font-size: 13.5px;

            background-color: #ffffff;

            outline: none;

            transition: all 0.2s ease;

            box-sizing: border-box;

            font-family: var(--font-family);

        }}

        .form-input:focus {{

            border-color: #3b82f6;

            box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.15);

        }}

        .table-row-toggle-spacer {{

            display: inline-block;

            width: 18px;

            text-align: center;

            margin-right: 8px;

            color: #9ca3af;

            font-weight: bold;

        }}

        /* Table Styling */

        .table-container {{

            border: 1px solid var(--border-color);

            border-radius: 8px;

            overflow: hidden;

            background-color: var(--bg-color);

            box-shadow: 0 1px 3px rgba(0,0,0,0.05);

            margin-top: 12px;

        }}

        .call-list-table {{

            width: 100%;

            border-collapse: collapse;

            text-align: left;

            font-size: 13.5px;

        }}

        .call-list-table th {{

            background-color: var(--surface-color);

            color: var(--text-color);

            font-weight: 600;

            padding: 12px 16px;

            border-bottom: 2px solid var(--border-color);

            font-size: 13px;

        }}

        .call-list-table td {{

            padding: 14px 16px;

            border-bottom: 1px solid var(--border-color);

            vertical-align: top;

        }}

        .call-list-table tbody tr:hover {{

            background-color: #fafbfc;

        }}

        /* Table priority badges */

        .priority-badge {{

            display: inline-block;

            font-size: 11px;

            font-weight: 700;

            padding: 3px 10px;

            border-radius: 12px;

            text-transform: uppercase;

            letter-spacing: 0.5px;

            text-align: center;

        }}

        .badge-relationship {{

            background-color: #dcfce7;

            color: #14532d;

            border: 1px solid #16a34a;

        }}

        .badge-high {{

            background-color: #fee2e2;

            color: #b91c1c;

            border: 1px solid #ef4444;

        }}

        .badge-medium {{

            background-color: #fef9c3;

            color: #a16207;

            border: 1px solid #eab308;

        }}

        .badge-low {{

            background-color: #f3f4f6;

            color: #374151;

            border: 1px solid #9ca3af;

        }}

        .col-name {{

            font-weight: 700;

            color: var(--text-color);

            font-size: 14px;

        }}

        .col-title {{

            font-size: 12px;

            color: var(--muted-color);

            margin-top: 2px;

        }}

        .col-org {{

            font-size: 10px;

            font-weight: 700;

            color: #475569;

            margin-top: 4px;

            text-transform: uppercase;

        }}

        .col-products {{

            font-weight: 500;

            color: #2563eb;

            font-size: 12px;

        }}

        .col-owner {{

            font-weight: 600;

            color: var(--text-color);

            font-size: 12.5px;

        }}

        .col-notes {{

            font-size: 12.5px;

            line-height: 1.4;

            color: #374151;

            max-width: 320px;

        }}

        .col-schedule {{

            font-weight: 700;

            font-size: 12px;

            color: #16a34a;

        }}

        /* Legend Panel */

        .legend-panel {{

            display: flex;

            gap: 16px;

            background-color: var(--surface-color);

            border: 1px solid var(--border-color);

            border-radius: 8px;

            padding: 12px 16px;

            margin-bottom: 24px;

            align-items: center;

            flex-wrap: wrap;

        }}

        .legend-title {{

            font-weight: 600;

            font-size: 13px;

            color: var(--muted-color);

            text-transform: uppercase;

            letter-spacing: 0.5px;

            margin-right: 8px;

        }}

        .legend-item {{

            display: flex;

            align-items: center;

            gap: 8px;

            font-size: 13px;

            font-weight: 500;

        }}

        .legend-color {{

            width: 14px;

            height: 14px;

            border-radius: 4px;

            border: 1px solid transparent;

        }}

        .lg-relationship {{

            background-color: var(--color-relationship-bg);

            border-color: var(--color-relationship-border);

        }}

        .lg-high {{

            background-color: var(--color-high-bg);

            border-color: var(--color-high-border);

        }}

        .lg-medium {{

            background-color: var(--color-medium-bg);

            border-color: var(--color-medium-border);

        }}

        .lg-low {{

            background-color: var(--color-low-bg);

            border-color: var(--color-low-border);

        }}

        /* Modal / Detail Panel */

        .detail-panel {{

            position: fixed;

            right: -400px;

            top: 0;

            width: 380px;

            height: 100%;

            background-color: var(--bg-color);

            border-left: 1px solid var(--border-color);

            box-shadow: -4px 0 15px rgba(0,0,0,0.08);

            z-index: 100;

            padding: 24px;

            overflow-y: auto;

            transition: right 0.3s ease;

        }}

        .detail-panel.active {{

            right: 0;

        }}

        .detail-header {{

            display: flex;

            justify-content: space-between;

            align-items: flex-start;

            margin-bottom: 20px;

            border-bottom: 1px solid var(--border-color);

            padding-bottom: 12px;

        }}

        .detail-close {{

            background: none;

            border: none;

            font-size: 20px;

            cursor: pointer;

            color: var(--muted-color);

        }}

        .detail-close:hover {{

            color: var(--text-color);

        }}

        .detail-title {{

            font-size: 18px;

            font-weight: 700;

            color: #1a365d;

        }}

        .detail-row {{

            margin-bottom: 16px;

        }}

        .detail-label {{

            font-size: 11px;

            font-weight: 700;

            color: var(--muted-color);

            text-transform: uppercase;

            letter-spacing: 0.5px;

            margin-bottom: 4px;

        }}

        .detail-value {{

            font-size: 13.5px;

            color: var(--text-color);

            background-color: var(--surface-color);

            padding: 10px;

            border-radius: 6px;

            border: 1px solid var(--border-color);

        }}

        .detail-value.products {{

            font-weight: 600;

            color: #2563eb;

            background-color: #f0f7ff;

            border-color: #bfdbfe;

        }}

        .detail-value.notes {{

            line-height: 1.5;

        }}

        .detail-overlay {{

            position: fixed;

            top: 0;

            left: 0;

            width: 100%;

            height: 100%;

            background-color: rgba(0,0,0,0.15);

            z-index: 99;

            display: none;

        }}

        .detail-overlay.active {{

            display: block;

        }}

        /* Made with IBM Bob footer */

        .bob-footer {{

            margin-top: 48px;

            border-top: 1px solid var(--border-color);

            padding-top: 16px;

            font-size: 12px;

            color: var(--muted-color);

            text-align: center;

        }}

        /* Backlog items or collapsed elements transition class */

        .collapsed-children {{

            display: none !important;

        }}

    </style>

</head>

<body>

    <div class="container">

        <header style="display: flex; justify-content: space-between; align-items: flex-start; flex-wrap: wrap; gap: 16px;">

            <div>

                <h1>SAP US Account Priority Technical Outreach Hub</h1>

            <div class="subtitle">v16 Rebuilt Core Directory | 376 Captured Contacts | 6 Main Org Branches + 1 Unchanged Branch</div>

            </div>

            <div style="display: flex; align-items: center; gap: 8px; font-weight: 600; font-size: 14px; color: #1a365d; background-color: var(--surface-color); padding: 8px 12px; border-radius: 6px; border: 1px solid var(--border-color);">

                <input type="checkbox" id="prioritized-only-checkbox" onchange="togglePrioritizedOnly(this.checked)" style="width: 16px; height: 16px; cursor: pointer; margin: 0;">

                <label for="prioritized-only-checkbox" style="cursor: pointer; user-select: none;">Only Show Prioritized</label>

            </div>

        </header>

        <!-- Legend Panel -->

        <div class="legend-panel">

            <span class="legend-title">Legend / Priority</span>

            <div class="legend-item">

                <div class="legend-color lg-relationship"></div>

                <span>Relationship (Direct / Green)</span>

            </div>

            <div class="legend-item">

                <div class="legend-color lg-high"></div>

                <span>High Priority Outreach (Red)</span>

            </div>

            <div class="legend-item">

                <div class="legend-color lg-medium"></div>

                <span>Medium Priority Outreach (Yellow)</span>

            </div>

            <div class="legend-item">

                <div class="legend-color lg-low"></div>

                <span>Low / Background Priority (Dark Grey)</span>

            </div>

        </div>

        <!-- Navigation Tabs -->

        <div class="tabs">

            <button class="tab-btn active" onclick="switchTab('org-chart-tab')">Interactive Org Chart (Family Tree)</button>

            <button class="tab-btn" onclick="switchTab('call-list-tab')">Priority Call List</button>

            <button class="tab-btn" onclick="switchTab('call-schedule-tab')">Call Schedule</button>

        </div>

        <!-- SEARCH / GLOBAL CONTROLS -->

        <div class="controls-row">

            <div class="search-container">

                <input type="text" id="global-search" class="search-input" placeholder="Search contact names or titles..." oninput="handleSearch(this.value)">

            </div>

            <div style="display: flex; align-items: center; gap: 6px;">

                <select id="seller-filter" class="form-input" onchange="handleSellerFilter(this.value)" style="min-width: 200px; height: 36px; font-size: 13px;">

                    <option value="">All Sellers</option>

                </select>

                <button class="btn" onclick="addNewSeller()" title="Add a new seller" style="height: 36px; white-space: nowrap;">+ Add Seller</button>

                <button class="btn" onclick="removeSellerGlobal()" title="Remove selected seller from all contacts" style="height: 36px; white-space: nowrap; background-color: #fef2f2; color: #b91c1c; border-color: #fca5a5;">X Remove Seller</button>

            </div>

            <div class="btn-group">

                <button class="btn" onclick="expandAllNodes()">Expand All</button>

                <button class="btn" onclick="collapseAllNodes()">Collapse All</button>

                <button class="btn primary" onclick="resetView()">Reset View</button>

            </div>

            <div class="btn-group" style="margin-left: auto;">

                <button class="btn" onclick="clearLocalStorage()" title="Clear any saved browser data and reload fresh" style="background:#fef2f2; color:#b91c1c; border-color:#fca5a5; font-size:12px;">Clear Saved Data</button>

                <button class="btn" onclick="downloadSchedule()" title="Permanently save all your edits into a new HTML file. Open that file next time to keep your changes." style="background:#1a365d; color:#fff; border-color:#1a365d; font-weight:700; font-size:13px; padding: 0 18px;">Save My Edits</button>

                <button id="gh-sync-btn" class="btn" onclick="openGitHubSyncPanel()" title="Sync your edits directly to GitHub so they save permanently for everyone" style="background:#24292f; color:#fff; border-color:#24292f; font-weight:700; font-size:13px; padding: 0 18px; display: flex; align-items: center; gap: 6px;">
                    <svg height="14" width="14" viewBox="0 0 16 16" fill="currentColor" style="vertical-align: middle;"><path d="M8 0C3.58 0 0 3.58 0 8c0 3.54 2.29 6.53 5.47 7.59.4.07.55-.17.55-.38 0-.19-.01-.82-.01-1.49-2.01.37-2.53-.49-2.69-.94-.09-.23-.48-.94-.82-1.13-.28-.15-.68-.52-.01-.53.63-.01 1.08.58 1.23.82.72 1.21 1.87.87 2.33.66.07-.52.28-.87.51-1.07-1.78-.2-3.64-.89-3.64-3.95 0-.87.31-1.59.82-2.15-.08-.2-.36-1.02.08-2.12 0 0 .67-.21 2.2.82.64-.18 1.32-.27 2-.27.68 0 1.36.09 2 .27 1.53-1.04 2.2-.82 2.2-.82.44 1.1.16 1.92.08 2.12.51.56.82 1.27.82 2.15 0 3.07-1.87 3.75-3.65 3.95.29.25.54.73.54 1.48 0 1.07-.01 1.93-.01 2.2 0 .21.15.46.55.38A8.012 8.012 0 0 0 16 8c0-4.42-3.58-8-8-8z"></path></svg>
                    GitHub Sync
                </button>

            </div>

        </div>

        <!-- TAB 1: ORG CHART -->

        <div id="org-chart-tab" class="tab-content active">

            <div class="org-tree-wrapper" id="org-tree-wrapper">

                <div class="org-tree-root" id="org-tree-root-container">

                    <!-- Tree will be rendered here dynamically -->

                </div>

                <!-- Floating Zoom and Pan Controls Panel -->

                <div class="zoom-controls">

                    <button class="zoom-btn" onclick="zoomIn()" title="Zoom In">+</button>

                    <button class="zoom-btn" onclick="zoomOut()" title="Zoom Out">-</button>

                    <button class="zoom-btn" onclick="zoomReset()" title="Reset / Center Fit">o</button>

                    <div class="zoom-info" id="zoom-percent">85%</div>

                </div>

            </div>

        </div>

        <!-- TAB 2: CALL LIST -->

        <div id="call-list-tab" class="tab-content">

            <div class="table-container">

                <table class="call-list-table" id="call-list-table">

                    <thead>

                        <tr>

                            <th>Contact Name & Title</th>

                            <th>LOB / Domain</th>

                            <th>Priority</th>

                            <th>Outreach Owner</th>

                            <th>Suggested Products</th>

                            <th>Why They Matter (Tech Fit)</th>

                            <th>Schedule</th>

                        </tr>

                    </thead>

                    <tbody id="call-list-tbody">

                        <!-- Table rows will be rendered here dynamically -->

                    </tbody>

                </table>

            </div>

        </div>

        <!-- TAB 3: CALL SCHEDULE -->

        <div id="call-schedule-tab" class="tab-content">

            <!-- Deadline banner -->

            <div id="deadline-banner" style="background: #1a365d; color: #fff; padding: 8px 16px; font-size: 12px; font-weight: 600; display: flex; align-items: center; gap: 16px; flex-wrap: wrap;">

                <span>Deal Deadline: <strong>Week of Dec 15, 2026</strong></span>

                <span id="deadline-countdown" style="background: rgba(255,255,255,0.15); padding: 2px 10px; border-radius: 99px;"></span>

                <span style="color: rgba(255,255,255,0.7); font-weight: 400;">Priority rules: Red = contact within 1 week | Yellow = within 5 weeks | Grey = within 15 weeks</span>

            </div>

            <!-- Toolbar -->

            <div style="display: flex; align-items: center; gap: 10px; padding: 12px 16px 10px; border-bottom: 1px solid #e5e7eb; background: #f7f8fa; flex-wrap: wrap;">

                <label for="schedule-seller-filter" style="font-size: 13px; font-weight: 600; color: #374151; white-space: nowrap;">Show schedule for:</label>

                <select id="schedule-seller-filter" class="form-input" onchange="renderScheduleTab()" style="min-width: 220px; height: 36px; font-size: 13px;">

                    <option value="">" Select a Seller "</option>

                </select>

                <select id="schedule-view-filter" class="form-input" onchange="renderScheduleTab()" style="height: 36px; font-size: 13px;">

                    <option value="all">All Weeks</option>

                    <option value="thisweek">This Week</option>

                    <option value="next2">Next 2 Weeks</option>

                    <option value="next4">Next 4 Weeks</option>

                    <option value="overdue">Overdue</option>

                </select>

                <span id="schedule-count-badge" style="font-size: 12px; color: #57606a; white-space: nowrap;"></span>

            </div>

            <!-- Schedule rendered as grouped week sections -->

            <div id="schedule-sections" style="padding: 12px 16px;">

                <p style="color: #57606a; font-size: 13px; text-align: center; padding: 32px 0;">Select a seller above to view their call schedule.</p>

            </div>

        </div>

        <!-- GitHub Sync Side Panel -->

        <div class="detail-overlay" id="github-sync-overlay" onclick="closeGitHubSyncPanel()"></div>

        <div class="detail-panel" id="github-sync-panel" style="width: 420px; right: -440px;">

            <div class="detail-header">

                <div class="detail-title" style="display: flex; align-items: center; gap: 8px; color: #24292f;">
                    <svg height="20" width="20" viewBox="0 0 16 16" fill="currentColor"><path d="M8 0C3.58 0 0 3.58 0 8c0 3.54 2.29 6.53 5.47 7.59.4.07.55-.17.55-.38 0-.19-.01-.82-.01-1.49-2.01.37-2.53-.49-2.69-.94-.09-.23-.48-.94-.82-1.13-.28-.15-.68-.52-.01-.53.63-.01 1.08.58 1.23.82.72 1.21 1.87.87 2.33.66.07-.52.28-.87.51-1.07-1.78-.2-3.64-.89-3.64-3.95 0-.87.31-1.59.82-2.15-.08-.2-.36-1.02.08-2.12 0 0 .67-.21 2.2.82.64-.18 1.32-.27 2-.27.68 0 1.36.09 2 .27 1.53-1.04 2.2-.82 2.2-.82.44 1.1.16 1.92.08 2.12.51.56.82 1.27.82 2.15 0 3.07-1.87 3.75-3.65 3.95.29.25.54.73.54 1.48 0 1.07-.01 1.93-.01 2.2 0 .21.15.46.55.38A8.012 8.012 0 0 0 16 8c0-4.42-3.58-8-8-8z"></path></svg>
                    GitHub Repository Sync
                </div>

                <button class="detail-close" onclick="closeGitHubSyncPanel()">&times;</button>

            </div>

            <div class="detail-row">
                <p style="font-size: 13px; color: var(--muted-color); line-height: 1.5; margin-bottom: 12px;">
                    Sync your edits directly back to your GitHub repository in real-time. This updates the file in your repository, which automatically triggers a deploy to GitHub Pages so everyone sees the updated portal.
                </p>
            </div>

            <div class="detail-row">

                <div class="detail-label">GitHub Personal Access Token (PAT)</div>

                <input type="password" id="gh-pat" placeholder="ghp_..." class="form-input" onchange="saveSyncCredentials()"/>

                <p style="font-size: 11.5px; color: var(--muted-color); margin-top: 6px; line-height: 1.4;">
                    Generate a token on GitHub with <b>repo</b> scope. Stored securely only in your own browser's local storage.
                    <a href="https://github.com/settings/tokens/new?scopes=repo&description=SAP-Org-Chart-Sync" target="_blank" style="color: #2563eb; text-decoration: underline;">Generate Token ↗</a>
                </p>

            </div>

            <div class="detail-row" style="display: grid; grid-template-columns: 1fr 1fr; gap: 12px;">
                <div>
                    <div class="detail-label">Repository Owner</div>
                    <input type="text" id="gh-owner" value="jreezycalhigh" class="form-input" onchange="saveSyncCredentials()"/>
                </div>
                <div>
                    <div class="detail-label">Repository Name</div>
                    <input type="text" id="gh-repo" value="SAP-Org-Chart" class="form-input" onchange="saveSyncCredentials()"/>
                </div>
            </div>

            <div class="detail-row" style="display: grid; grid-template-columns: 1.2fr 0.8fr; gap: 12px;">
                <div>
                    <div class="detail-label">File Path</div>
                    <input type="text" id="gh-path" value="SAP_Account_Priority_Outreach_Hub.html" class="form-input" onchange="saveSyncCredentials()"/>
                </div>
                <div>
                    <div class="detail-label">Branch</div>
                    <input type="text" id="gh-branch" value="main" class="form-input" onchange="saveSyncCredentials()"/>
                </div>
            </div>

            <div class="detail-row" style="margin-top: 24px; border-top: 1px solid var(--border-color); padding-top: 16px;">
                <div id="gh-sync-status" style="font-size: 12.5px; color: var(--muted-color); margin-bottom: 12px; padding: 10px; border-radius: 6px; background-color: var(--surface-color); border: 1px solid var(--border-color); min-height: 20px; display: flex; align-items: center; justify-content: center; text-align: center; line-height: 1.4;">
                    Ready to sync.
                </div>
                
                <button class="btn primary" onclick="syncEditsToGitHub()" style="width: 100%; justify-content: center; height: 42px; font-weight: 700; background: #24292f; border-color: #24292f; color: #fff; cursor: pointer; display: flex; align-items: center; gap: 8px;">
                    <svg height="16" width="16" viewBox="0 0 16 16" fill="currentColor"><path d="M8 0C3.58 0 0 3.58 0 8c0 3.54 2.29 6.53 5.47 7.59.4.07.55-.17.55-.38 0-.19-.01-.82-.01-1.49-2.01.37-2.53-.49-2.69-.94-.09-.23-.48-.94-.82-1.13-.28-.15-.68-.52-.01-.53.63-.01 1.08.58 1.23.82.72 1.21 1.87.87 2.33.66.07-.52.28-.87.51-1.07-1.78-.2-3.64-.89-3.64-3.95 0-.87.31-1.59.82-2.15-.08-.2-.36-1.02.08-2.12 0 0 .67-.21 2.2.82.64-.18 1.32-.27 2-.27.68 0 1.36.09 2 .27 1.53-1.04 2.2-.82 2.2-.82.44 1.1.16 1.92.08 2.12.51.56.82 1.27.82 2.15 0 3.07-1.87 3.75-3.65 3.95.29.25.54.73.54 1.48 0 1.07-.01 1.93-.01 2.2 0 .21.15.46.55.38A8.012 8.012 0 0 0 16 8c0-4.42-3.58-8-8-8z"></path></svg>
                    Commit & Push to GitHub
                </button>
            </div>

        </div>

        <!-- Detail Side Panel -->

        <div class="detail-overlay" id="detail-overlay" onclick="closeDetailPanel()"></div>

        <div class="detail-panel" id="detail-panel">

            <div class="detail-header">

                <div class="detail-title" id="det-name">Contact Details</div>

                <button class="detail-close" onclick="closeDetailPanel()">&times;</button>

            </div>

            <!-- View Mode Container -->

            <div id="view-mode-container">

                <div class="detail-row">

                    <div class="detail-label">Title</div>

                    <div class="detail-value" id="det-title">-</div>

                </div>

                <div class="detail-row">

                    <div class="detail-label">Line of Business / Domain</div>

                    <div class="detail-value" id="det-org">-</div>

                </div>

                <div class="detail-row">

                    <div class="detail-label">Priority</div>

                    <div class="detail-value" id="det-priority">-</div>

                </div>

                <div class="detail-row">

                    <div class="detail-label">IBM Outreach Owner (Seller)</div>

                    <div class="detail-value" id="det-owner">-</div>

                </div>

                <div class="detail-row">

                    <div class="detail-label">Suggested IBM Products</div>

                    <div class="detail-value products" id="det-products">-</div>

                </div>

                <div class="detail-row">

                    <div class="detail-label">Technology Fit & Notes</div>

                    <div class="detail-value notes" id="det-notes">-</div>

                </div>

                <div class="detail-row">

                    <div class="detail-label">Public Verification (LinkedIn, ZoomInfo, Web)</div>

                    <div class="detail-value" id="det-verification" style="font-weight: 600; color: #16a34a;">-</div>

                </div>

                <div class="detail-row">

                    <div class="detail-label">Verification Notes / Role Insights</div>

                    <div class="detail-value notes" id="det-verification-note" style="line-height: 1.4; color: #374151;">-</div>

                </div>

                <div class="detail-row">

                    <div class="detail-label">Call Schedule</div>

                    <div class="detail-value" id="det-schedule">-</div>

                </div>

                <button class="btn primary" onclick="enterEditMode()" style="width: 100%; margin-top: 12px; justify-content: center;">Edit Contact Details</button>

            </div>

            <!-- Edit Mode Container (Hidden by default) -->

            <div id="edit-mode-container" style="display: none;">

                <div class="detail-row">

                    <div class="detail-label">Title</div>

                    <input type="text" id="edit-title" class="form-input">

                </div>

                <div class="detail-row">

                    <div class="detail-label">Priority</div>

                    <select id="edit-priority" class="form-input">

                        <option value="Relationship">Relationship</option>

                        <option value="High">High</option>

                        <option value="Medium">Medium</option>

                        <option value="Low">Low</option>

                    </select>

                </div>

                <div class="detail-row">

                    <div class="detail-label">IBM Outreach Owner (Seller)</div>

                    <div style="display: flex; gap: 6px;">

                        <select id="edit-owner" class="form-input" style="flex: 1;"></select>

                        <button class="btn" onclick="addNewSellerFromEdit()" style="white-space: nowrap; height: 36px;">+ New</button>

                        <button class="btn" onclick="removeSellerFromEdit()" style="white-space: nowrap; height: 36px; background-color: #fef2f2; color: #b91c1c; border-color: #fca5a5;">. Remove</button>

                    </div>

                </div>

                <div class="detail-row">

                    <div class="detail-label">Suggested IBM Products</div>

                    <input type="text" id="edit-products" class="form-input" placeholder="e.g. watsonx Orchestrate, IBM Bob">

                </div>

                <div class="detail-row">

                    <div class="detail-label">Call Schedule</div>

                    <select id="edit-schedule" class="form-input">

                        <option value="N/A">N/A</option>

                        {" ".join(f'<option value="Week {{w}}">Week {{w}}</option>' for w in range(1, 31))}

                    </select>

                </div>

                <div class="detail-row">

                    <div class="detail-label">Technology Fit & Notes</div>

                    <textarea id="edit-notes" class="form-input" style="height: 100px; resize: vertical;"></textarea>

                </div>

                <div style="display: flex; gap: 8px; margin-top: 16px;">

                    <button class="btn primary" onclick="saveContactEdits()" style="flex: 1; justify-content: center;">Save</button>

                    <button class="btn" onclick="cancelContactEdits()" style="flex: 1; justify-content: center;">Cancel</button>

                </div>

                <button class="btn" onclick="deleteContact()" style="width: 100%; margin-top: 12px; background-color: #fef2f2; color: #b91c1c; border-color: #fca5a5; justify-content: center;">Delete Contact</button>

            </div>

        </div>

        <div class="bob-footer">

            Made with IBM Bob

        </div>

    </div>

    <!-- Data Injection -->

    <script>

        let orgData = {json.dumps(root_node, ensure_ascii=False)};

        let callListData = {json.dumps(call_list, ensure_ascii=False)};

        // Filter state

        let onlyShowPrioritized = false;

        let currentContactName = null;

        let sellerFilter = "";

        // Priority row background colors for call list

        const priorityRowColors = {{

            "relationship": "#f0fdf4",

            "high":         "#fef2f2",

            "medium":       "#fefce8",

            "low":          "#f9fafb"

        }};

        const priorityRowBorders = {{

            "relationship": "#16a34a",

            "high":         "#ef4444",

            "medium":       "#eab308",

            "low":          "#e5e7eb"

        }};

        function applyRowColor(tr, priority) {{

            const p = (priority || "low").toLowerCase();

            tr.style.backgroundColor = priorityRowColors[p] || "#f9fafb";

            tr.style.borderLeft = `4px solid ${{priorityRowBorders[p] || "#e5e7eb"}}`;

        }}

        // -- Seller utilities --------------------------------------------------

        // Split owner strings on "/" only when outside parentheses, so

        // "Ryan Sorber (IBM Data/AI/Watson)" stays as one token.

        function splitOwners(ownerStr) {{

            const parts = [];

            let depth = 0, cur = "";

            for (let i = 0; i < ownerStr.length; i++) {{

                const ch = ownerStr[i];

                if (ch === "(") {{ depth++; cur += ch; }}

                else if (ch === ")") {{ depth--; cur += ch; }}

                else if (ch === "/" && depth === 0) {{ parts.push(cur.trim()); cur = ""; }}

                else {{ cur += ch; }}

            }}

            if (cur.trim()) parts.push(cur.trim());

            return parts;

        }}

        // Strip trailing parenthetical suffix: "Jerome Carlson (ATL)" ' "Jerome Carlson"

        function normaliseSellerName(raw) {{

            return raw.replace(/ *[(][^)]*[)]/g, "").trim();

        }}

        // Extra pool for sellers added at runtime

        var _extraSellerPool = [];

        // Return deduplicated, normalised, sorted seller list

        function getAllSellers() {{

            const seen = new Set();

            const result = [];

            function addSeller(raw) {{

                const norm = normaliseSellerName(raw);

                if (norm && !seen.has(norm.toLowerCase())) {{

                    seen.add(norm.toLowerCase());

                    result.push(norm);

                }}

            }}

            function collectSellers(node) {{

                if (node.owner) splitOwners(node.owner).forEach(s => addSeller(s));

                (node.children || []).forEach(collectSellers);

            }}

            collectSellers(orgData);

            callListData.forEach(c => {{ if (c.owner) splitOwners(c.owner).forEach(s => addSeller(s)); }});

            _extraSellerPool.forEach(s => addSeller(s));

            return result.filter(Boolean).sort();

        }}

        function populateSellerDropdowns(selectedOwner) {{

            const sellers = getAllSellers();

            // Top filter bar dropdown

            const filterSelect = document.getElementById("seller-filter");

            const filterCurrent = filterSelect.value;

            filterSelect.innerHTML = '<option value="">All Sellers</option>';

            sellers.forEach(s => {{

                const opt = document.createElement("option");

                opt.value = s; opt.textContent = s;

                if (s === filterCurrent) opt.selected = true;

                filterSelect.appendChild(opt);

            }});

            // Call Schedule tab seller dropdown

            const schedSel = document.getElementById("schedule-seller-filter");

            if (schedSel) {{

                const schedCurrent = schedSel.value;

                schedSel.innerHTML = '<option value="">" Select a Seller "</option>';

                sellers.forEach(s => {{

                    const opt = document.createElement("option");

                    opt.value = s; opt.textContent = s;

                    if (s === schedCurrent) opt.selected = true;

                    schedSel.appendChild(opt);

                }});

            }}

            // Edit panel owner dropdown

            const editSelect = document.getElementById("edit-owner");

            if (editSelect) {{

                editSelect.innerHTML = "";

                sellers.forEach(s => {{

                    const opt = document.createElement("option");

                    opt.value = s; opt.textContent = s;

                    if (selectedOwner && s.toLowerCase() === selectedOwner.toLowerCase()) opt.selected = true;

                    editSelect.appendChild(opt);

                }});

            }}

        }}

        function handleSellerFilter(val) {{

            sellerFilter = val;

            renderOrgTree();

            renderCallList();

        }}

        function addNewSeller() {{

            const name = prompt("Enter new seller name:");

            if (!name || !name.trim()) return;

            const trimmed = name.trim();

            // Add them as a dummy entry so they appear in the dropdown

            // They become real once assigned to a contact via edit

            if (!getAllSellers().includes(trimmed)) {{

                // Temporarily inject into a callListData entry's owner so it persists in the set

                if (callListData.length > 0) {{

                    callListData[0]._extraSellers = callListData[0]._extraSellers || [];

                    callListData[0]._extraSellers.push(trimmed);

                    // Patch getAllSellers to also read _extraSellers

                }}

                _extraSellerPool.push(trimmed);

            }}

            populateSellerDropdowns(trimmed);

            document.getElementById("seller-filter").value = trimmed;

            handleSellerFilter(trimmed);

        }}

        function addNewSellerFromEdit() {{

            const name = prompt("Enter new seller name:");

            if (!name || !name.trim()) return;

            const trimmed = name.trim();

            if (!_extraSellerPool.includes(trimmed)) _extraSellerPool.push(trimmed);

            populateSellerDropdowns(trimmed);

            document.getElementById("edit-owner").value = trimmed;

        }}

        function removeSellerFromEdit() {{

            const sel = document.getElementById("edit-owner");

            const sellerToRemove = sel.value;

            if (!sellerToRemove) return;

            if (!confirm(`Remove "${{sellerToRemove}}" from the owner dropdown? Any contact currently assigned to this seller will be reassigned to Jerome Carlson (ATL).`)) return;

            // Reassign all contacts that had this seller

            const fallback = "Jerome Carlson (ATL)";

            function reassignTree(node) {{

                if ((node.owner || "").split("/").map(s=>s.trim()).includes(sellerToRemove)) {{

                    node.owner = fallback;

                }}

                (node.children || []).forEach(reassignTree);

            }}

            reassignTree(orgData);

            callListData.forEach(c => {{

                if ((c.owner || "").split("/").map(s=>s.trim()).includes(sellerToRemove)) c.owner = fallback;

            }});

            // Remove from pool if present

            _extraSellerPool = _extraSellerPool.filter(s => s !== sellerToRemove);

            populateSellerDropdowns(fallback);

            renderOrgTree();

            renderCallList();

        }}

        function removeSellerGlobal() {{

            const sel = document.getElementById("seller-filter");

            const sellerToRemove = sel.value;

            if (!sellerToRemove) {{ alert("Select a seller in the dropdown first to remove them."); return; }}

            if (!confirm(`Remove "${{sellerToRemove}}" from the system? All contacts assigned to this seller will be reassigned to Jerome Carlson (ATL).`)) return;

            const fallback = "Jerome Carlson (ATL)";

            function reassignTree(node) {{

                if ((node.owner || "").split("/").map(s=>s.trim()).includes(sellerToRemove)) {{

                    node.owner = fallback;

                }}

                (node.children || []).forEach(reassignTree);

            }}

            reassignTree(orgData);

            callListData.forEach(c => {{

                if ((c.owner || "").split("/").map(s=>s.trim()).includes(sellerToRemove)) c.owner = fallback;

            }});

            _extraSellerPool = _extraSellerPool.filter(s => s !== sellerToRemove);

            populateSellerDropdowns();

            sel.value = "";

            handleSellerFilter("");

        }}

        function matchesSeller(owner) {{

            if (!sellerFilter) return true;

            return splitOwners(owner || "").map(s => normaliseSellerName(s)).includes(sellerFilter);

        }}

        function togglePrioritizedOnly(checked) {{

            onlyShowPrioritized = checked;

            renderOrgTree();

            renderCallList();

        }}

        function hasPrioritizedDescendantOrSelf(node) {{

            if (node.priority && node.priority !== "Low") return true;

            if (node.children) {{

                for (let child of node.children) {{

                    if (hasPrioritizedDescendantOrSelf(child)) return true;

                }}

            }}

            return false;

        }}

        // CRM Edit / Update logic

        function enterEditMode() {{

            document.getElementById("view-mode-container").style.display = "none";

            document.getElementById("edit-mode-container").style.display = "block";

            const title = document.getElementById("det-title").innerText;

            const priority = document.getElementById("det-priority").innerText;

            const owner = document.getElementById("det-owner").innerText;

            const schedule = document.getElementById("det-schedule").innerText;

            const notes = document.getElementById("det-notes").innerText;

            const products = document.getElementById("det-products").innerText;

            document.getElementById("edit-title").value = title;

            document.getElementById("edit-priority").value = priority;

            document.getElementById("edit-notes").value = notes;

            document.getElementById("edit-products").value = products === "-" ? "" : products;

            // Set schedule dropdown

            const schedSel = document.getElementById("edit-schedule");

            schedSel.value = schedule;

            if (!schedSel.value) schedSel.value = "N/A";

            // Populate seller dropdown and select current owner (normalise so it matches dropdown values)

            populateSellerDropdowns(normaliseSellerName(owner));

        }}

        function cancelContactEdits() {{

            document.getElementById("view-mode-container").style.display = "block";

            document.getElementById("edit-mode-container").style.display = "none";

        }}

        function updateNodeInTree(node, targetName, updatedFields) {{

            if (node.name.toLowerCase() === targetName.toLowerCase()) {{

                Object.assign(node, updatedFields);

                return true;

            }}

            if (node.children) {{

                for (let child of node.children) {{

                    if (updateNodeInTree(child, targetName, updatedFields)) {{

                        return true;

                    }}

                }}

            }}

            return false;

        }}

        function deleteNodeFromTree(node, targetName) {{

            if (node.children) {{

                const index = node.children.findIndex(child => child.name.toLowerCase() === targetName.toLowerCase());

                if (index !== -1) {{

                    node.children.splice(index, 1);

                    return true;

                }}

                for (let child of node.children) {{

                    if (deleteNodeFromTree(child, targetName)) {{

                        return true;

                    }}

                }}

            }}

            return false;

        }}

        function saveContactEdits() {{

            const newTitle = document.getElementById("edit-title").value.trim();

            const newPriority = document.getElementById("edit-priority").value;

            const newOwner = document.getElementById("edit-owner").value;

            let newSchedule = document.getElementById("edit-schedule").value;

            const newNotes = document.getElementById("edit-notes").value.trim();

            const newProducts = document.getElementById("edit-products").value.trim();

            const savedName = currentContactName;

            if (!newTitle) {{ alert("Job title cannot be empty!"); return; }}

            // Auto-recalculate schedule if priority changed to something more urgent

            const oldPriority = document.getElementById("det-priority").innerText;

            const tempContact = {{ priority: newPriority, schedule: newSchedule }};

            if (newPriority !== oldPriority) {{

                recalcSchedule(tempContact, true);

                newSchedule = tempContact.schedule;

            }} else {{

                recalcSchedule(tempContact, false);

                newSchedule = tempContact.schedule;

            }}

            // Update orgData tree

            updateNodeInTree(orgData, savedName, {{

                title: newTitle, priority: newPriority, owner: newOwner,

                schedule: newSchedule, notes: newNotes, products: newProducts

            }});

            // Update callListData in-place by index

            const idx = callListData.findIndex(c => c.name.toLowerCase() === savedName.toLowerCase());

            if (idx !== -1) {{

                callListData[idx].title    = newTitle;

                callListData[idx].priority = newPriority;

                callListData[idx].owner    = newOwner;

                callListData[idx].schedule = newSchedule;

                callListData[idx].notes    = newNotes;

                callListData[idx].products = newProducts;

            }}

            // Patch org-chart card in-place (tree stays expanded)

            const cardId = "card-" + savedName.toLowerCase().replace(/[^a-z0-9]/g, "-");

            const card = document.getElementById(cardId);

            if (card) {{

                card.className = card.className.replace(/priority-[^ ]+/, "priority-" + newPriority);

                const pill = card.querySelector(".card-priority-pill");

                if (pill) pill.innerText = newPriority;

                const titleEl = card.querySelector(".card-title");

                if (titleEl) titleEl.innerText = newTitle;

            }}

            // Close edit mode, write fields directly, keep panel open

            document.getElementById("edit-mode-container").style.display = "none";

            document.getElementById("view-mode-container").style.display = "block";

            document.getElementById("det-name").innerText     = savedName;

            document.getElementById("det-title").innerText    = newTitle;

            document.getElementById("det-priority").innerText = newPriority;

            document.getElementById("det-owner").innerText    = newOwner;

            document.getElementById("det-schedule").innerText = newSchedule;

            document.getElementById("det-notes").innerText    = newNotes;

            document.getElementById("det-products").innerText = newProducts || "-";

            // Refresh all three tabs so they reflect the change immediately

            renderCallList();

            renderScheduleTab();

            // Autosave to localStorage so edits survive page close/reopen

            saveToLocalStorage();

            // If GitHub Sync credentials are saved, auto-sync back to GitHub!
            const pat = localStorage.getItem(GH_KEY_PREFIX + "pat");
            if (pat && pat.length > 5) {{
                syncEditsToGitHubBackground();
            }}

        }}

         function deleteContact() {{

             if (confirm(`Are you sure you want to delete ${{currentContactName}} from the org chart?`)) {{

                 deleteNodeFromTree(orgData, currentContactName);

                 callListData = callListData.filter(c => c.name.toLowerCase() !== currentContactName.toLowerCase());

                 closeDetailPanel();

                 renderOrgTree();

                 renderCallList();

                 // Autosave to localStorage so edits survive page close/reopen
                 saveToLocalStorage();

                 // If GitHub Sync credentials are saved, auto-sync back to GitHub!
                 const pat = localStorage.getItem(GH_KEY_PREFIX + "pat");
                 if (pat && pat.length > 5) {{
                     syncEditsToGitHubBackground();
                 }}

             }}

         }}

        // Zoom and Pan State Variables

        let isDragging = false;

        let startX, startY;

        let panX = 0;

        let panY = 180;

        let zoom = 0.85;

        // Expanded nodes state for the table view

        const expandedTableNodes = new Set();

        function buildTableRows(node, depth, rowsArray) {{

            if (onlyShowPrioritized && !hasPrioritizedDescendantOrSelf(node)) {{

                return;

            }}

            rowsArray.push({{ node: node, depth: depth }});

            if (expandedTableNodes.has(node.name.toLowerCase())) {{

                if (node.children && node.children.length > 0) {{

                    const visibleChildren = node.children.filter(child => !onlyShowPrioritized || hasPrioritizedDescendantOrSelf(child));

                    visibleChildren.forEach(child => {{

                        buildTableRows(child, depth + 1, rowsArray);

                    }});

                }}

            }}

        }}

        function toggleTableNode(nameLower, event) {{

            if (expandedTableNodes.has(nameLower)) {{

                expandedTableNodes.delete(nameLower);

            }} else {{

                expandedTableNodes.add(nameLower);

            }}

            renderCallList();

        }}

        function updateTransform() {{

            const container = document.getElementById("org-tree-root-container");

            container.style.transform = `translate(calc(-50% + ${{panX}}px), ${{panY}}px) scale(${{zoom}})`;

            document.getElementById("zoom-percent").innerText = Math.round(zoom * 100) + "%";

        }}

        function zoomIn() {{

            zoom = Math.min(2.0, zoom + 0.03);

            updateTransform();

        }}

        function zoomOut() {{

            zoom = Math.max(0.15, zoom - 0.03);

            updateTransform();

        }}

        function zoomReset() {{

            zoom = 0.85;

            panX = 0;

            panY = 180;

            updateTransform();

        }}

        function initPanZoom() {{

            const viewport = document.getElementById("org-tree-wrapper");

            viewport.addEventListener("mousedown", (e) => {{

                if (e.target.closest(".card-name-toggle") || e.target.closest("button") || e.target.closest(".node-card")) return;

                isDragging = true;

                viewport.style.cursor = "grabbing";

                startX = e.clientX - panX;

                startY = e.clientY - panY;

            }});

            window.addEventListener("mousemove", (e) => {{

                if (!isDragging) return;

                panX = e.clientX - startX;

                panY = e.clientY - startY;

                updateTransform();

            }});

            window.addEventListener("mouseup", () => {{

                isDragging = false;

                viewport.style.cursor = "grab";

            }});

            viewport.addEventListener("wheel", (e) => {{

                e.preventDefault();

                const zoomFactor = 0.015;

                if (e.deltaY < 0) {{

                    zoom = Math.min(2.0, zoom + zoomFactor);

                }} else {{

                    zoom = Math.max(0.15, zoom - zoomFactor);

                }}

                updateTransform();

            }});

        }}

        // -- localStorage Autosave --

        const LS_KEY = "sap_outreach_hub_edits";

        function saveToLocalStorage() {{

            try {{

                const edits = {{}};

                callListData.forEach(c => {{

                    const id = (c.name || "").toLowerCase().replace(/[^a-z0-9]/g, "_");

                    edits[id] = {{

                        priority:      c.priority      || "",

                        owner:         c.owner         || "",

                        products:      c.products      || "",

                        notes:         c.notes         || "",

                        schedule:      c.schedule      || "",

                        schedule_week: c.schedule_week || 0,

                        title:         c.title         || ""

                    }};

                }});

                localStorage.setItem(LS_KEY, JSON.stringify(edits));

            }} catch(e) {{ console.warn("Autosave failed:", e.message); }}

        }}

        function clearLocalStorage() {{
            localStorage.removeItem(LS_KEY);
            alert("Local saved data cleared. Reloading...");
            location.reload();
        }}

        function loadFromLocalStorage() {{

            try {{

                const raw = localStorage.getItem(LS_KEY);

                if (!raw) return;

                let edits;
                try {{ edits = JSON.parse(raw); }} catch(e) {{
                    console.warn("Corrupt localStorage — clearing");
                    localStorage.removeItem(LS_KEY);
                    return;
                }}

                // Apply to orgData tree

                function applyToTree(node) {{

                    if (!node) return;

                    const id = (node.name || "").toLowerCase().replace(/[^a-z0-9]/g, "_");

                    const e = edits[id];

                    if (e) {{

                        if (e.priority)      node.priority      = e.priority;

                        if (e.owner)         node.owner         = e.owner;

                        if (e.products)      node.products      = e.products;

                        if (e.notes)         node.notes         = e.notes;

                        if (e.schedule)      node.schedule      = e.schedule;

                        if (e.schedule_week) node.schedule_week = e.schedule_week;

                        if (e.title)         node.title         = e.title;

                    }}

                    if (node.children) node.children.forEach(c => applyToTree(c));

                }}

                applyToTree(orgData);

                // Apply to callListData

                callListData.forEach(c => {{

                    const id = (c.name || "").toLowerCase().replace(/[^a-z0-9]/g, "_");

                    const e = edits[id];

                    if (e) {{

                        if (e.priority)      c.priority      = e.priority;

                        if (e.owner)         c.owner         = e.owner;

                        if (e.products)      c.products      = e.products;

                        if (e.notes)         c.notes         = e.notes;

                        if (e.schedule)      c.schedule      = e.schedule;

                        if (e.schedule_week) c.schedule_week = e.schedule_week;

                        if (e.title)         c.title         = e.title;

                    }}

                }});

            }} catch(e) {{ console.warn("Load from storage failed:", e.message); }}

        }}

        // -- End localStorage Autosave --

        // Initialize App

        document.addEventListener("DOMContentLoaded", () => {{

            loadFromLocalStorage();

            populateSellerDropdowns();

            renderOrgTree();

            renderCallList();

            initPanZoom();

            updateTransform();

        }});

        function switchTab(tabId) {{

            document.querySelectorAll(".tab-content").forEach(el => el.classList.remove("active"));

            document.querySelectorAll(".tab-btn").forEach(el => el.classList.remove("active"));

            document.getElementById(tabId).classList.add("active");

            event.target.classList.add("active");

            if (tabId === "call-list-tab") {{

                document.getElementById("global-search").placeholder = "Filter call list by name, title, products...";

            }} else if (tabId === "call-schedule-tab") {{

                document.getElementById("global-search").placeholder = "Search contact names or titles...";

                renderScheduleTab();

            }} else {{

                document.getElementById("global-search").placeholder = "Search contact names or titles...";

            }}

        }}

        // Render Org Tree

        function renderOrgTree() {{

            const rootContainer = document.getElementById("org-tree-root-container");

            rootContainer.innerHTML = "";

            const treeHTML = buildTreeNodeHTML(orgData);

            rootContainer.appendChild(treeHTML);

            // Collapse everything on initial render

            document.querySelectorAll("#org-tree-root-container .tree-children").forEach(el => el.classList.add("collapsed-children"));

            document.querySelectorAll("#org-tree-root-container .card-name-toggle").forEach(el => el.innerText = "+");

            updateAllWrappingClasses();

        }}

        function buildTreeNodeHTML(node) {{

            const container = document.createElement("div");

            container.className = "tree-node-container";

            container.id = "node-container-" + node.name.toLowerCase().replace(/[^a-z0-9]/g, "-");

            // Card element

            const card = document.createElement("div");

            card.className = `node-card priority-${{node.priority}}`;

            card.id = "card-" + node.name.toLowerCase().replace(/[^a-z0-9]/g, "-");

            card.setAttribute("onclick", `showContactDetails(${{JSON.stringify(node.name)}})`);

            // Priority pill

            const pill = document.createElement("span");

            pill.className = "card-priority-pill";

            pill.innerText = node.priority;

            card.appendChild(pill);

            // Name

            const nameEl = document.createElement("div");

            nameEl.className = "card-name";

            nameEl.innerText = node.name + " ";

            let cardToggleBtn = null;

            const currentVisibleChildren = node.children ? node.children.filter(child => !onlyShowPrioritized || hasPrioritizedDescendantOrSelf(child)) : [];

            if (currentVisibleChildren.length > 0) {{

                cardToggleBtn = document.createElement("span");

                cardToggleBtn.className = "card-name-toggle";

                cardToggleBtn.innerText = "+"; // default is collapsed

                cardToggleBtn.onclick = (e) => {{

                    e.stopPropagation(); // prevent modal trigger

                    toggleNodeChildren(container, cardToggleBtn);

                }};

                nameEl.appendChild(cardToggleBtn);

            }}

            card.appendChild(nameEl);

            // Title

            const titleEl = document.createElement("div");

            titleEl.className = "card-title";

            titleEl.innerText = node.title || "Executive";

            card.appendChild(titleEl);

            // Org

            const orgEl = document.createElement("div");

            orgEl.className = "card-org";

            orgEl.innerText = node.org || "";

            card.appendChild(orgEl);

            container.appendChild(card);

            // Sibling elements / children

            if (currentVisibleChildren.length > 0) {{

                // Connector line

                const connector = document.createElement("div");

                connector.className = "tree-node-card-connector";

                container.appendChild(connector);

                // Group adjacent children sharing the same LOB into visual dotted boxes!

                const childrenWrapper = document.createElement("div");

                childrenWrapper.className = "tree-children";

                let currentLOBGroup = [];

                let currentLOB = null;

                const primaryLOBs = ["Ariba", "NS2", "CDX", "SuccessFactors", "Concur", "CX", "Sovereign Cloud", "Business Network"];

                const lobDisplayNames = {{
                    "Ariba":            "Ariba — Procurement & Supply Chain",
                    "NS2":              "NS2 — US Public Sector",
                    "CDX":              "CDX — Cloud DevOps & Operations",
                    "SuccessFactors":   "SuccessFactors — HCM Suite",
                    "Concur":           "Concur — Travel & Expense",
                    "CX":               "Cloud Products & Customer Experience",
                    "Sovereign Cloud":  "Sovereign Cloud & Infrastructure",
                    "Business Network": "Business Network"
                }};

                function renderLOBGroup(lobName, items) {{

                    if (lobName) {{

                        const groupDiv = document.createElement("div");

                        const lobClass = lobName.replace(/[^a-zA-Z0-9]/g, "");

                        groupDiv.className = `lob-group-box lob-${{lobClass}}`;

                        const titleLabel = document.createElement("div");

                        titleLabel.className = "lob-group-box-title";

                        titleLabel.innerText = lobDisplayNames[lobName] || lobName;

                        groupDiv.appendChild(titleLabel);

                        items.forEach(child => {{

                            const childHTML = buildTreeNodeHTML(child);

                            groupDiv.appendChild(childHTML);

                        }});

                        childrenWrapper.appendChild(groupDiv);

                    }} else {{

                        items.forEach(child => {{

                            const childHTML = buildTreeNodeHTML(child);

                            childrenWrapper.appendChild(childHTML);

                        }});

                    }}

                }}

                currentVisibleChildren.forEach(child => {{

                    const childLOB = child.org || "";

                    let matchedLOB = null;

                    for (let p of primaryLOBs) {{

                        if (childLOB.toLowerCase().includes(p.toLowerCase())) {{

                            matchedLOB = p;

                            break;

                        }}

                    }}

                    if (matchedLOB && (currentLOB === null || matchedLOB === currentLOB)) {{

                        currentLOBGroup.push(child);

                        currentLOB = matchedLOB;

                    }} else {{

                        if (currentLOBGroup.length > 0) {{

                            renderLOBGroup(currentLOB, currentLOBGroup);

                        }}

                        currentLOBGroup = [child];

                        currentLOB = matchedLOB;

                    }}

                }});

                if (currentLOBGroup.length > 0) {{

                    renderLOBGroup(currentLOB, currentLOBGroup);

                }}

                container.appendChild(childrenWrapper);

            }}

            return container;

        }}

        function updateAllWrappingClasses() {{

            // Loop through all .tree-children containers

            document.querySelectorAll(".tree-children").forEach(tc => {{

                let hasExpandedChild = false;

                // Direct children of this .tree-children that are .tree-node-container

                const directContainers = tc.querySelectorAll(":scope > .tree-node-container");

                directContainers.forEach(container => {{

                    const subChildren = container.querySelector(":scope > .tree-children");

                    if (subChildren && !subChildren.classList.contains("collapsed-children")) {{

                        hasExpandedChild = true;

                    }}

                }});

                // Children inside any .lob-group-box under this .tree-children

                const lobGroupBoxes = tc.querySelectorAll(":scope > .lob-group-box");

                lobGroupBoxes.forEach(box => {{

                    const boxContainers = box.querySelectorAll(":scope > .tree-node-container");

                    boxContainers.forEach(container => {{

                        const subChildren = container.querySelector(":scope > .tree-children");

                        if (subChildren && !subChildren.classList.contains("collapsed-children")) {{

                            hasExpandedChild = true;

                        }}

                    }});

                }});

                if (hasExpandedChild) {{

                    tc.classList.add("has-expanded-child");

                }} else {{

                    tc.classList.remove("has-expanded-child");

                }}

            }});

            // Loop through all .lob-group-box containers

            document.querySelectorAll(".lob-group-box").forEach(box => {{

                let hasExpandedChild = false;

                const boxContainers = box.querySelectorAll(":scope > .tree-node-container");

                boxContainers.forEach(container => {{

                    const subChildren = container.querySelector(":scope > .tree-children");

                    if (subChildren && !subChildren.classList.contains("collapsed-children")) {{

                        hasExpandedChild = true;

                    }}

                }});

                if (hasExpandedChild) {{

                    box.classList.add("has-expanded-child");

                }} else {{

                    box.classList.remove("has-expanded-child");

                }}

            }});

        }}

        function toggleNodeChildren(container, btn) {{

            const childrenWrapper = container.querySelector(":scope > .tree-children");

            if (childrenWrapper) {{

                if (childrenWrapper.classList.contains("collapsed-children")) {{

                    childrenWrapper.classList.remove("collapsed-children");

                    if (btn) btn.innerText = "-";

                }} else {{

                    childrenWrapper.classList.add("collapsed-children");

                    if (btn) btn.innerText = "+";

                }}

                updateAllWrappingClasses();

            }}

        }}

        function expandAllNodes() {{

            document.querySelectorAll(".tree-children").forEach(el => el.classList.remove("collapsed-children"));

            document.querySelectorAll(".card-name-toggle").forEach(el => el.innerText = "-");

            updateAllWrappingClasses();

        }}

        function collapseAllNodes() {{

            document.querySelectorAll(".tree-children").forEach(el => el.classList.add("collapsed-children"));

            document.querySelectorAll(".card-name-toggle").forEach(el => el.innerText = "+");

            updateAllWrappingClasses();

        }}

        // -- Schedule Engine --

        // Deadline: week of Dec 15 2026. "Week 1" = week of today.

        // Priority ' target week window from NOW:

        //   Relationship ' N/A (no schedule)

        //   High (Red)   ' within 1 week of NOW

        //   Medium (Yellow) ' within 5 weeks of NOW

        //   Low (Grey)   ' within 15 weeks of NOW

        const DEADLINE = new Date("2026-12-15");

        function getWeekStart(d) {{

            const dt = new Date(d);

            const day = dt.getDay(); // 0=Sun

            dt.setDate(dt.getDate() - day);

            dt.setHours(0,0,0,0);

            return dt;

        }}

        function getCurrentWeekNumber() {{

            // Returns the ISO-style "week number" relative to the start of this year

            // We use a simpler model: week 1 = this calendar week, week 2 = next, etc.

            return 1; // always "this week" for relative calculations

        }}

        function weeksBetween(d1, d2) {{

            return Math.round((d2 - d1) / (7 * 24 * 3600 * 1000));

        }}

        function scheduleWeekToDate(weekStr) {{

            // "Week N" ' actual date of that Monday, counting from this week

            const m = (weekStr || "").match(/Week[ \t]+([0-9]+)/i);

            if (!m) return null;

            const w = parseInt(m[1]);

            const base = getWeekStart(new Date());

            const d = new Date(base);

            d.setDate(d.getDate() + (w - 1) * 7);

            return d;

        }}

        function dateToWeekStr(d) {{

            const now = getWeekStart(new Date());

            const diff = weeksBetween(now, getWeekStart(d));

            if (diff < 0) return "Overdue";

            if (diff === 0) return "This Week";

            return "Week " + (diff + 1);

        }}

        function priorityToMaxWeeks(priority) {{

            if (priority === "Relationship") return null; // N/A

            if (priority === "High") return 1;

            if (priority === "Medium") return 5;

            return 15;

        }}

        // Recalculate a contact's schedule based on current priority, preserving

        // manual overrides unless priority changed to a more urgent bucket.

        function recalcSchedule(contact, forceRecalc) {{

            const p = contact.priority;

            if (p === "Relationship") {{ contact.schedule = "N/A"; return; }}

            const maxW = priorityToMaxWeeks(p);

            const current = (contact.schedule || "");

            const mw = current.match(/Week[ \t]+([0-9]+)/i);

            const currentW = mw ? parseInt(mw[1]) : 999;

            // If priority changed to something more urgent than current slot, push earlier

            if (forceRecalc || currentW > maxW || current === "Backlog" || current === "N/A") {{

                // Assign to next available week within the priority window

                const deadline = getWeekStart(DEADLINE);

                const now = getWeekStart(new Date());

                const weeksLeft = weeksBetween(now, deadline);

                // Spread within the window, don't go past deadline

                const targetWeek = Math.min(maxW, weeksLeft);

                contact.schedule = targetWeek <= 0 ? "This Week" : "Week " + targetWeek;

            }}

        }}

        // Update deadline countdown

        function updateDeadlineCountdown() {{

            const el = document.getElementById("deadline-countdown");

            if (!el) return;

            const now = new Date();

            const diff = weeksBetween(getWeekStart(now), getWeekStart(DEADLINE));

            if (diff <= 0) {{

                el.textContent = "  DEADLINE PASSED";

                el.style.background = "rgba(220,38,38,0.6)";

            }} else {{

                el.textContent = diff + " weeks remaining";

                el.style.background = diff <= 4 ? "rgba(220,38,38,0.4)" : "rgba(255,255,255,0.15)";

            }}

        }}

        updateDeadlineCountdown();

        // -- Call Schedule Tab --

        function renderScheduleTab() {{

            const seller = document.getElementById("schedule-seller-filter").value;

            const viewFilter = (document.getElementById("schedule-view-filter") || {{}}).value || "all";

            const container = document.getElementById("schedule-sections");

            const badge = document.getElementById("schedule-count-badge");

            container.innerHTML = "";

            if (!seller) {{

                container.innerHTML = '<p style="color:#57606a;font-size:13px;text-align:center;padding:32px 0;">Select a seller above to view their call schedule.</p>';

                badge.textContent = "";

                return;

            }}

            // Gather contacts for this seller

            const sellerNorm = seller.toLowerCase();

            const matched = [];

            function collectForSeller(node) {{

                const ownerParts = splitOwners(node.owner || "").map(s => normaliseSellerName(s).toLowerCase());

                if (ownerParts.some(o => o === sellerNorm)) matched.push(node);

                (node.children || []).forEach(collectForSeller);

            }}

            collectForSeller(orgData);

            callListData.forEach(c => {{

                const ownerParts = splitOwners(c.owner || "").map(s => normaliseSellerName(s).toLowerCase());

                if (ownerParts.some(o => o === sellerNorm)) {{

                    if (!matched.find(m => m.name.toLowerCase() === c.name.toLowerCase())) matched.push(c);

                }}

            }});

            // Determine current week number relative to today

            const nowWS = getWeekStart(new Date());

            const deadlineWS = getWeekStart(DEADLINE);

            const totalWeeks = weeksBetween(nowWS, deadlineWS);

            // Compute actual calendar date for each contact's schedule slot

            function getContactWeekNum(c) {{

                const s = c.schedule || "";

                if (s === "N/A") return -1;

                if (s === "This Week" || s === "Overdue") return 0;

                const m = s.match(/Week[ \t]+([0-9]+)/i);

                return m ? parseInt(m[1]) : 999;

            }}

            function getContactDate(c) {{

                const w = getContactWeekNum(c);

                if (w < 0) return null;

                const d = new Date(nowWS);

                d.setDate(d.getDate() + (Math.max(0, w - 1)) * 7);

                return d;

            }}

            // Apply view filter

            let filtered = matched.filter(c => {{

                const w = getContactWeekNum(c);

                if (viewFilter === "all") return true;

                if (viewFilter === "thisweek") return w === 0 || w === 1;

                if (viewFilter === "next2") return w >= 0 && w <= 2;

                if (viewFilter === "next4") return w >= 0 && w <= 4;

                if (viewFilter === "overdue") return w === 0 && (c.schedule || "").toLowerCase().includes("overdue");

                return true;

            }});

            badge.textContent = filtered.length + " contact" + (filtered.length !== 1 ? "s" : "");

            if (filtered.length === 0) {{

                container.innerHTML = '<p style="color:#57606a;font-size:13px;text-align:center;padding:32px 0;">No contacts match this filter.</p>';

                return;

            }}

            // Sort by week number then priority

            const prioOrder = {{"Relationship": 0, "High": 1, "Medium": 2, "Low": 3}};

            filtered.sort((a, b) => {{

                const wa = getContactWeekNum(a), wb = getContactWeekNum(b);

                if (wa !== wb) return wa - wb;

                return (prioOrder[a.priority] ?? 9) - (prioOrder[b.priority] ?? 9);

            }});

            // Group by week label

            const groups = {{}};

            filtered.forEach(c => {{

                const w = getContactWeekNum(c);

                let label;

                let dateLabel = "";

                if (w < 0) {{

                    label = "N/A - Relationship Contacts";

                }} else if (w === 0 || c.schedule === "This Week") {{

                    label = " This Week";

                    dateLabel = nowWS.toLocaleDateString("en-US", {{month:"short", day:"numeric"}}) + " - " + new Date(nowWS.getTime()+6*86400000).toLocaleDateString("en-US", {{month:"short", day:"numeric"}});

                }} else {{

                    const d = getContactDate(c);

                    const end = d ? new Date(d.getTime()+6*86400000) : null;

                    label = "Week " + w;

                    if (d && d <= deadlineWS) {{

                        dateLabel = d.toLocaleDateString("en-US", {{month:"short", day:"numeric"}}) + " - " + (end ? end.toLocaleDateString("en-US", {{month:"short", day:"numeric"}}) : "");

                        if (d > deadlineWS) label += "   Past Deadline";

                    }} else if (d && d > deadlineWS) {{

                        label += "   Past Deadline";

                    }}

                }}

                if (!groups[label]) groups[label] = {{dateLabel, items: []}};

                groups[label].items.push(c);

            }});

            // Render each group

            const urgentStyle = "border-left: 3px solid #dc2626;";

            const thisWeekStyle = "border-left: 3px solid #2563eb; background: #eff6ff;";

            Object.entries(groups).forEach(([label, group]) => {{

                const isThisWeek = label.startsWith("");

                const isPast = label.includes("  Past Deadline");

                const isNA = label.startsWith("N/A");

                // Section header

                const hdr = document.createElement("div");

                hdr.style.cssText = `display:flex;align-items:baseline;gap:10px;margin:18px 0 6px;padding-bottom:6px;border-bottom:2px solid ${{isThisWeek ? "#2563eb" : isPast ? "#dc2626" : isNA ? "#9ca3af" : "#e5e7eb"}};`;

                hdr.innerHTML = `

                    <span style="font-size:15px;font-weight:700;color:${{isThisWeek?"#1d4ed8":isPast?"#dc2626":isNA?"#9ca3af":"#1f2328"}}">${{label}}</span>

                    ${{group.dateLabel ? `<span style="font-size:12px;color:#57606a;">${{group.dateLabel}}</span>` : ""}}

                    <span style="font-size:11px;color:#9ca3af;margin-left:auto">${{group.items.length}} contact${{group.items.length!==1?"s":""}}</span>

                `;

                container.appendChild(hdr);

                // Table for this group

                const tbl = document.createElement("table");

                tbl.className = "call-list-table";

                tbl.style.marginBottom = "4px";

                tbl.innerHTML = `<thead><tr>

                    <th style="width:200px">Contact</th>

                    <th>LOB</th>

                    <th style="width:90px">Priority</th>

                    <th style="width:130px">Products</th>

                    <th>Why They Matter</th>

                    <th style="width:80px">Deadline</th>

                </tr></thead>`;

                const tbody = document.createElement("tbody");

                group.items.forEach(c => {{

                    const tr = document.createElement("tr");

                    tr.style.cursor = "pointer";

                    tr.setAttribute("onclick", `showContactDetails(${{JSON.stringify(c.name)}})`);

                    applyRowColor(tr, c.priority);

                    if (isThisWeek) tr.style.cssText += "border-left:3px solid #2563eb;";

                    // Weeks until deadline

                    const cDate = getContactDate(c);

                    const weeksUntil = cDate ? weeksBetween(nowWS, getWeekStart(cDate)) : null;

                    const deadlineCell = weeksUntil === null ? "N/A"

                        : weeksUntil === 0 ? '<span style="color:#dc2626;font-weight:700">This Week</span>'

                        : weeksUntil < 0 ? '<span style="color:#dc2626;font-weight:700">Overdue</span>'

                        : `<span style="color:#374151">${{weeksUntil}}w left</span>`;

                    tr.innerHTML = `

                        <td><div class="col-name">${{c.name}}</div><div class="col-title">${{c.title||""}}</div></td>

                        <td><div class="col-org">${{c.org||""}}</div></td>

                        <td><span class="priority-badge badge-${{(c.priority||"low").toLowerCase()}}">${{c.priority||"Low"}}</span></td>

                        <td><div class="col-products" style="font-size:11px">${{c.products||""}}</div></td>

                        <td><div class="col-notes">${{c.notes||""}}</div></td>

                        <td style="text-align:center">${{deadlineCell}}</td>

                    `;

                    tbody.appendChild(tr);

                }});

                tbl.appendChild(tbody);

                container.appendChild(tbl);

            }});

        }}

        // Download / share the current state as a standalone HTML file

        function downloadSchedule() {{

            const seller = document.getElementById("schedule-seller-filter").value || "All Sellers";

            const content = document.documentElement.outerHTML;

            const blob = new Blob([content], {{type: "text/html"}});

            const url = URL.createObjectURL(blob);

            const a = document.createElement("a");

            a.href = url;

            a.download = "SAP_Outreach_Hub_" + seller.replace(/[^a-zA-Z0-9]/g,"_") + "_" + new Date().toISOString().slice(0,10) + ".html";

            a.click();

            URL.revokeObjectURL(url);

        }}

        function resetView() {{

            document.getElementById("global-search").value = "";

            document.querySelectorAll(".node-card").forEach(c => c.classList.remove("highlighted"));

            collapseAllNodes();

            expandedTableNodes.clear();

            renderCallList();

            zoomReset();

        }}

        // Render Call List Table

        function renderCallList() {{

            const tbody = document.getElementById("call-list-tbody");

            tbody.innerHTML = "";

            const searchVal = document.getElementById("global-search").value.toLowerCase().trim();

            const isSearchActive = searchVal.length >= 2;

            if (isSearchActive) {{

                // Flat filtered list sorted by priority

                let filtered = callListData.filter(c => {{

                    const text = (c.name + " " + c.title + " " + c.org + " " + c.products + " " + c.notes).toLowerCase();

                    return text.includes(searchVal);

                }});

                if (onlyShowPrioritized) {{

                    filtered = filtered.filter(c => c.priority !== "Low");

                }}

                if (sellerFilter) {{

                    filtered = filtered.filter(c => matchesSeller(c.owner));

                }}

                filtered.forEach(c => {{

                    const tr = document.createElement("tr");

                    tr.setAttribute("onclick", `showContactDetails(${{JSON.stringify(c.name)}})`);

                    tr.style.cursor = "pointer";

                    tr.id = "tr-" + c.name.toLowerCase().replace(/[^a-z0-9]/g, "-");

                    applyRowColor(tr, c.priority);

                    // Column 1: Name & Title (Flat - no indentation or row toggles)

                    const nameTd = document.createElement("td");

                    const nameDiv = document.createElement("div");

                    nameDiv.className = "col-name";

                    nameDiv.innerText = c.name;

                    const titleDiv = document.createElement("div");

                    titleDiv.className = "col-title";

                    titleDiv.innerText = c.title;

                    nameTd.appendChild(nameDiv);

                    nameTd.appendChild(titleDiv);

                    tr.appendChild(nameTd);

                    // Column 2: LOB

                    const orgTd = document.createElement("td");

                    const orgDiv = document.createElement("div");

                    orgDiv.className = "col-org";

                    orgDiv.innerText = c.org;

                    orgTd.appendChild(orgDiv);

                    tr.appendChild(orgTd);

                    // Column 3: Priority

                    const priTd = document.createElement("td");

                    const priSpan = document.createElement("span");

                    priSpan.className = `priority-badge badge-${{c.priority.toLowerCase()}}`;

                    priSpan.innerText = c.priority;

                    priTd.appendChild(priSpan);

                    tr.appendChild(priTd);

                    // Column 4: Owner

                    const ownerTd = document.createElement("td");

                    const ownerDiv = document.createElement("div");

                    ownerDiv.className = "col-owner";

                    ownerDiv.innerText = c.owner;

                    ownerTd.appendChild(ownerDiv);

                    tr.appendChild(ownerTd);

                    // Column 5: Products

                    const prodTd = document.createElement("td");

                    const prodDiv = document.createElement("div");

                    prodDiv.className = "col-products";

                    prodDiv.innerText = c.products;

                    prodTd.appendChild(prodDiv);

                    tr.appendChild(prodTd);

                    // Column 6: Notes

                    const notesTd = document.createElement("td");

                    const notesDiv = document.createElement("div");

                    notesDiv.className = "col-notes";

                    notesDiv.innerText = c.notes;

                    notesTd.appendChild(notesDiv);

                    tr.appendChild(notesTd);

                    // Column 7: Schedule

                    const schedTd = document.createElement("td");

                    const schedDiv = document.createElement("div");

                    schedDiv.className = "col-schedule";

                    schedDiv.innerText = c.schedule;

                    schedTd.appendChild(schedDiv);

                    tr.appendChild(schedTd);

                    tbody.appendChild(tr);

                }});

            }} else {{

                // Hierarchical folded list starting with only the CEO

                const rowsArray = [];

                buildTableRows(orgData, 0, rowsArray);

                rowsArray.filter(c => !sellerFilter || matchesSeller(c.node.owner)).forEach(c => {{

                    const tr = document.createElement("tr");

                    tr.setAttribute("onclick", `showContactDetails(${{JSON.stringify(c.node.name)}})`);

                    tr.style.cursor = "pointer";

                    tr.id = "tr-" + c.node.name.toLowerCase().replace(/[^a-z0-9]/g, "-");

                    applyRowColor(tr, c.node.priority);

                    // Column 1: Name & Title (Hierarchical - with indentation & row toggles)

                    const nameTd = document.createElement("td");

                    nameTd.style.paddingLeft = `${{c.depth * 24 + 16}}px`;

                    const nameLine = document.createElement("div");

                    nameLine.className = "col-name";

                    if (c.node.children && c.node.children.length > 0) {{

                        const btn = document.createElement("button");

                        btn.className = "table-row-toggle-btn";

                        btn.innerText = expandedTableNodes.has(c.node.name.toLowerCase()) ? "-" : "+";

                        btn.onclick = (e) => {{

                            e.stopPropagation();

                            toggleTableNode(c.node.name.toLowerCase(), e);

                        }};

                        nameLine.appendChild(btn);

                    }} else {{

                        const spacer = document.createElement("span");

                        spacer.className = "table-row-toggle-spacer";

                        spacer.innerText = "";

                        nameLine.appendChild(spacer);

                    }}

                    const textSpan = document.createElement("span");

                    textSpan.innerText = c.node.name;

                    nameLine.appendChild(textSpan);

                    nameTd.appendChild(nameLine);

                    const titleDiv = document.createElement("div");

                    titleDiv.className = "col-title";

                    titleDiv.style.paddingLeft = (c.node.children && c.node.children.length > 0) ? "26px" : "18px";

                    titleDiv.innerText = c.node.title;

                    nameTd.appendChild(titleDiv);

                    tr.appendChild(nameTd);

                    // Column 2: LOB

                    const orgTd = document.createElement("td");

                    const orgDiv = document.createElement("div");

                    orgDiv.className = "col-org";

                    orgDiv.innerText = c.node.org || "";

                    orgTd.appendChild(orgDiv);

                    tr.appendChild(orgTd);

                    // Column 3: Priority

                    const priTd = document.createElement("td");

                    const priSpan = document.createElement("span");

                    priSpan.className = `priority-badge badge-${{c.node.priority.toLowerCase()}}`;

                    priSpan.innerText = c.node.priority;

                    priTd.appendChild(priSpan);

                    tr.appendChild(priTd);

                    // Column 4: Owner

                    const ownerTd = document.createElement("td");

                    const ownerDiv = document.createElement("div");

                    ownerDiv.className = "col-owner";

                    ownerDiv.innerText = c.node.owner || "Jerome Carlson (ATL)";

                    ownerTd.appendChild(ownerDiv);

                    tr.appendChild(ownerTd);

                    // Column 5: Products

                    const prodTd = document.createElement("td");

                    const prodDiv = document.createElement("div");

                    prodDiv.className = "col-products";

                    prodDiv.innerText = c.node.products || "";

                    prodTd.appendChild(prodDiv);

                    tr.appendChild(prodTd);

                    // Column 6: Notes

                    const notesTd = document.createElement("td");

                    const notesDiv = document.createElement("div");

                    notesDiv.className = "col-notes";

                    notesDiv.innerText = c.node.notes || "";

                    notesTd.appendChild(notesDiv);

                    tr.appendChild(notesTd);

                    // Column 7: Schedule

                    const schedTd = document.createElement("td");

                    const schedDiv = document.createElement("div");

                    schedDiv.className = "col-schedule";

                    schedDiv.innerText = c.node.schedule || "";

                    schedTd.appendChild(schedDiv);

                    tr.appendChild(schedTd);

                    tbody.appendChild(tr);

                }});

            }}

        }}

        // Search functionality

        function handleSearch(val) {{

            const searchVal = val.toLowerCase().trim();

            // Tab 1 search: Org Chart

            const allCards = document.querySelectorAll(".node-card");

            allCards.forEach(c => c.classList.remove("highlighted"));

            if (searchVal.length >= 2) {{

                // Find matching names

                const matches = [];

                // Recurse function to search names in orgData

                function searchInTree(node, parents = []) {{

                    if (node.name.toLowerCase().includes(searchVal) || node.title.toLowerCase().includes(searchVal)) {{

                        matches.push({{ name: node.name, parents: [...parents] }});

                    }}

                    if (node.children) {{

                        const visibleChildren = node.children.filter(child => !onlyShowPrioritized || hasPrioritizedDescendantOrSelf(child));

                        visibleChildren.forEach(child => {{

                            searchInTree(child, [...parents, node.name]);

                        }});

                    }}

                }}

                searchInTree(orgData);

                // Highlight and auto-expand matches!

                matches.forEach(m => {{

                    const cardId = "card-" + m.name.toLowerCase().replace(/[^a-z0-9]/g, "-");

                    const card = document.getElementById(cardId);

                    if (card) {{

                        card.classList.add("highlighted");

                        // Expand parents so this match is visible!

                        m.parents.forEach(p => {{

                            const parentId = "node-container-" + p.toLowerCase().replace(/[^a-z0-9]/g, "-");

                            const parentContainer = document.getElementById(parentId);

                            if (parentContainer) {{

                                const childrenWrapper = parentContainer.querySelector(".tree-children");

                                const btn = parentContainer.querySelector(".card-name-toggle");

                                const connector = parentContainer.querySelector(".tree-node-card-connector");

                                if (childrenWrapper && childrenWrapper.classList.contains("collapsed-children")) {{

                                    childrenWrapper.classList.remove("collapsed-children");

                                    if (connector) connector.style.display = "block";

                                    if (btn) btn.innerText = "-";

                                }}

                            }}

                        }});

                    }}

                }});

                updateAllWrappingClasses();

            }} else {{

                updateAllWrappingClasses();

            }}

            // Tab 2 search: Call List (filtering table rows)

            renderCallList();

        }}

        // Modal Display Contact Details

        function showContactDetails(name) {{

            cancelContactEdits();

            const lowerName = name.toLowerCase();

            currentContactName = name;

            // Find contact details in either contacts or build a temporary one for Manoj / branch heads

            let contact = callListData.find(c => c.name.toLowerCase() === lowerName);

            if (!contact) {{

                // Try to find in OrgTree (might be Manoj or Branch Head)

                function findInTree(node) {{

                    if (node.name.toLowerCase() === lowerName) return node;

                    if (node.children) {{

                        for (let child of node.children) {{

                            const found = findInTree(child);

                            if (found) return found;

                        }}

                    }}

                    return null;

                }}

                const found = findInTree(orgData);

                if (found) {{

                    contact = {{

                        name: found.name,

                        title: found.title,

                        org: found.org,

                        priority: found.priority,

                        owner: found.owner || "Jerome Carlson (ATL)",

                        products: found.products || "watsonx.governance",

                        notes: found.notes || "Branch Head or Executive.",

                        schedule: found.schedule || "N/A",

                        verification: found.verification || "Verified",

                        verification_note: found.verification_note || "Branch Head node."

                    }};

                }}

            }}

            if (contact) {{

                document.getElementById("det-name").innerText = contact.name;

                document.getElementById("det-title").innerText = contact.title;

                document.getElementById("det-org").innerText = contact.org;

                document.getElementById("det-priority").innerText = contact.priority;

                document.getElementById("det-owner").innerText = contact.owner;

                document.getElementById("det-products").innerText = contact.products;

                document.getElementById("det-notes").innerText = contact.notes;

                document.getElementById("det-schedule").innerText = contact.schedule;

                const verEl = document.getElementById("det-verification");

                verEl.innerText = contact.verification || "Not yet run through the public-verification pass.";

                if (verEl.innerText === "Verified") {{

                    verEl.style.color = "#16a34a";

                }} else if (verEl.innerText === "Not verified") {{

                    verEl.style.color = "#ef4444";

                }} else {{

                    verEl.style.color = "#57606a";

                }}

                document.getElementById("det-verification-note").innerText = contact.verification_note || "No verification notes available.";

                // Show panel

                document.getElementById("detail-panel").classList.add("active");

                document.getElementById("detail-overlay").classList.add("active");

            }}

        }}

        function closeDetailPanel() {{

            document.getElementById("detail-panel").classList.remove("active");

            document.getElementById("detail-overlay").classList.remove("active");

        }}

        // -- GitHub Repository Sync Integration --

        const GH_KEY_PREFIX = "sap_gh_";

        function loadSyncCredentials() {{
            try {{
                const pat = localStorage.getItem(GH_KEY_PREFIX + "pat") || "";
                const owner = localStorage.getItem(GH_KEY_PREFIX + "owner") || "jreezycalhigh";
                const repo = localStorage.getItem(GH_KEY_PREFIX + "repo") || "SAP-Org-Chart";
                const path = localStorage.getItem(GH_KEY_PREFIX + "path") || "SAP_Account_Priority_Outreach_Hub.html";
                const branch = localStorage.getItem(GH_KEY_PREFIX + "branch") || "main";

                document.getElementById("gh-pat").value = pat;
                document.getElementById("gh-owner").value = owner;
                document.getElementById("gh-repo").value = repo;
                document.getElementById("gh-path").value = path;
                document.getElementById("gh-branch").value = branch;

                updateGitHubSyncButtonState();
            }} catch (e) {{
                console.warn("Failed to load GitHub credentials from localStorage:", e);
            }}
        }}

        function saveSyncCredentials() {{
            try {{
                localStorage.setItem(GH_KEY_PREFIX + "pat", document.getElementById("gh-pat").value.trim());
                localStorage.setItem(GH_KEY_PREFIX + "owner", document.getElementById("gh-owner").value.trim());
                localStorage.setItem(GH_KEY_PREFIX + "repo", document.getElementById("gh-repo").value.trim());
                localStorage.setItem(GH_KEY_PREFIX + "path", document.getElementById("gh-path").value.trim());
                localStorage.setItem(GH_KEY_PREFIX + "branch", document.getElementById("gh-branch").value.trim());
                
                updateGitHubSyncButtonState();
            }} catch (e) {{
                console.warn("Failed to save GitHub credentials to localStorage:", e);
            }}
        }}

        function updateGitHubSyncButtonState() {{
            const pat = localStorage.getItem(GH_KEY_PREFIX + "pat");
            const btn = document.getElementById("gh-sync-btn");
            if (btn) {{
                if (pat && pat.length > 5) {{
                    btn.style.background = "#15803d"; // Green when configured
                    btn.style.borderColor = "#15803d";
                    btn.title = "GitHub Sync is configured. Click to push changes.";
                }} else {{
                    btn.style.background = "#24292f"; // Dark when not configured
                    btn.style.borderColor = "#24292f";
                    btn.title = "Click to set up GitHub Sync and push changes.";
                }}
            }}
        }}

        function openGitHubSyncPanel() {{
            // Close details first
            closeDetailPanel();
            
            // Load latest saved credentials
            loadSyncCredentials();

            // Set active
            document.getElementById("github-sync-panel").classList.add("active");
            document.getElementById("github-sync-overlay").classList.add("active");
        }}

        function closeGitHubSyncPanel() {{
            document.getElementById("github-sync-panel").classList.remove("active");
            document.getElementById("github-sync-overlay").classList.remove("active");
        }}

        async function syncEditsToGitHub() {{
            const statusEl = document.getElementById("gh-sync-status");
            const pat = document.getElementById("gh-pat").value.trim();
            const owner = document.getElementById("gh-owner").value.trim();
            const repo = document.getElementById("gh-repo").value.trim();
            const path = document.getElementById("gh-path").value.trim();
            const branch = document.getElementById("gh-branch").value.trim();

            if (!pat) {{
                statusEl.innerHTML = '<span style="color:#dc2626; font-weight:700;">Error: GitHub Personal Access Token (PAT) is required!</span>';
                return;
            }}

            statusEl.innerHTML = '<span style="color:#2563eb; font-weight:700;">⏳ Connecting to GitHub...</span>';

            try {{
                // Step 1: Fetch the file metadata to get SHA and original content
                const url = `https://api.github.com/repos/${{owner}}/${{repo}}/contents/${{path}}?ref=${{branch}}`;
                const getRes = await fetch(url, {{
                    headers: {{
                        "Authorization": `token ${{pat}}`,
                        "Accept": "application/vnd.github.v3+json",
                        "Cache-Control": "no-cache"
                    }}
                }});

                if (!getRes.ok) {{
                    const errText = await getRes.text();
                    throw new Error(`Failed to fetch file from GitHub (${{getRes.status}}): ${{errText}}`);
                }}

                const fileData = await getRes.json();
                const sha = fileData.sha;
                
                // Decode original content
                const b64Raw = fileData.content.replace(/\s/g, "");
                const originalHTML = decodeURIComponent(escape(atob(b64Raw)));

                statusEl.innerHTML = '<span style="color:#2563eb; font-weight:700;">🔄 Preparing updated dataset...</span>';

                // Step 2: Inject our updated in-memory dataset
                let updatedHTML = originalHTML;

                // Replace orgData line
                const orgDataRegex = /let orgData = \\\\{{[^]*?\\\\}}/;
                const newOrgDataLine = "let orgData = " + JSON.stringify(orgData) + ";";
                if (orgDataRegex.test(updatedHTML)) {{
                    updatedHTML = updatedHTML.replace(orgDataRegex, newOrgDataLine);
                }} else {{
                    const fallbackRegex = /let orgData = \\\\{{.*?\\\\}}/;
                    updatedHTML = updatedHTML.replace(fallbackRegex, "let orgData = " + JSON.stringify(orgData) + ";");
                }}

                // Replace callListData line
                const callListDataRegex = /let callListData = \\\\[[^]*?\\\\];/;
                const newCallListDataLine = "let callListData = " + JSON.stringify(callListData) + ";";
                if (callListDataRegex.test(updatedHTML)) {{
                    updatedHTML = updatedHTML.replace(callListDataRegex, newCallListDataLine);
                }} else {{
                    const fallbackRegex = /let callListData = \\\\[.*?\\\\];/;
                    updatedHTML = updatedHTML.replace(fallbackRegex, "let callListData = " + JSON.stringify(callListData) + ";");
                }}

                statusEl.innerHTML = '<span style="color:#2563eb; font-weight:700;">📤 Committing and pushing to GitHub...</span>';

                // UTF-8 base64 encoding helper
                const encodedContent = btoa(unescape(encodeURIComponent(updatedHTML)));

                // Step 3: PUT back to GitHub to commit
                const putRes = await fetch(`https://api.github.com/repos/${{owner}}/${{repo}}/contents/${{path}}`, {{
                    method: "PUT",
                    headers: {{
                        "Authorization": `token ${{pat}}`,
                        "Accept": "application/vnd.github.v3+json",
                        "Content-Type": "application/json"
                    }},
                    body: JSON.stringify({{
                        message: "Update outreach contact list and org chart data (via Browser GUI Sync)",
                        content: encodedContent,
                        sha: sha,
                        branch: branch
                    }})
                }});

                if (!putRes.ok) {{
                    const errText = await putRes.text();
                    throw new Error(`Failed to save file to GitHub (${{putRes.status}}): ${{errText}}`);
                }}

                statusEl.innerHTML = '<span style="color:#16a34a; font-weight:700;">✅ Success! Changes committed & pushed directly.<br><span style="font-size:11px; font-weight:normal; color:#475569;">GitHub Pages will update your live portal in 30-60 seconds.</span></span>';
                
                // Clear any local localStorage edits since they are now pushed permanently to master!
                localStorage.removeItem(LS_KEY);
                
                updateGitHubSyncButtonState();

            }} catch (error) {{
                console.error("Sync Error:", error);
                statusEl.innerHTML = `<span style="color:#dc2626; font-weight:700;">Error: ${{error.message}}</span>`;
            }}
        }}

        function showToast(message, type = "info") {{
            let toast = document.getElementById("gh-toast");
            if (!toast) {{
                toast = document.createElement("div");
                toast.id = "gh-toast";
                toast.style.position = "fixed";
                toast.style.bottom = "24px";
                toast.style.left = "24px";
                toast.style.padding = "12px 18px";
                toast.style.borderRadius = "8px";
                toast.style.backgroundColor = "#24292f";
                toast.style.color = "#fff";
                toast.style.fontSize = "13px";
                toast.style.fontWeight = "600";
                toast.style.boxShadow = "0 4px 12px rgba(0,0,0,0.15)";
                toast.style.zIndex = "9999";
                toast.style.display = "flex";
                toast.style.alignItems = "center";
                toast.style.gap = "8px";
                toast.style.transition = "all 0.3s ease";
                document.body.appendChild(toast);
            }}
            
            if (type === "success") {{
                toast.style.backgroundColor = "#15803d";
                toast.innerHTML = '✅ ' + message;
            }} else if (type === "error") {{
                toast.style.backgroundColor = "#b91c1c";
                toast.innerHTML = '❌ ' + message;
            }} else {{
                toast.style.backgroundColor = "#24292f";
                toast.innerHTML = '⏳ ' + message;
            }}
            
            toast.style.opacity = "1";
            toast.style.transform = "translateY(0)";
            
            if (type !== "info") {{
                setTimeout(() => {{
                    toast.style.opacity = "0";
                    toast.style.transform = "translateY(10px)";
                }}, 4000);
            }}
        }}

        async function syncEditsToGitHubBackground() {{
            const pat = localStorage.getItem(GH_KEY_PREFIX + "pat") || "";
            const owner = localStorage.getItem(GH_KEY_PREFIX + "owner") || "jreezycalhigh";
            const repo = localStorage.getItem(GH_KEY_PREFIX + "repo") || "SAP-Org-Chart";
            const path = localStorage.getItem(GH_KEY_PREFIX + "path") || "SAP_Account_Priority_Outreach_Hub.html";
            const branch = localStorage.getItem(GH_KEY_PREFIX + "branch") || "main";

            if (!pat) return;

            showToast("Syncing your changes to GitHub...", "info");

            try {{
                const url = `https://api.github.com/repos/${{owner}}/${{repo}}/contents/${{path}}?ref=${{branch}}`;
                const getRes = await fetch(url, {{
                    headers: {{
                        "Authorization": `token ${{pat}}`,
                        "Accept": "application/vnd.github.v3+json",
                        "Cache-Control": "no-cache"
                    }}
                }});

                if (!getRes.ok) {{
                    throw new Error(`Fetch failed (${{getRes.status}})`);
                }}

                const fileData = await getRes.json();
                const sha = fileData.sha;
                
                const b64Raw = fileData.content.replace(/\s/g, "");
                const originalHTML = decodeURIComponent(escape(atob(b64Raw)));

                let updatedHTML = originalHTML;

                // Replace orgData line
                const orgDataRegex = /let orgData = \\\\{{[^]*?\\\\}}/;
                const newOrgDataLine = "let orgData = " + JSON.stringify(orgData) + ";";
                if (orgDataRegex.test(updatedHTML)) {{
                    updatedHTML = updatedHTML.replace(orgDataRegex, newOrgDataLine);
                }} else {{
                    const fallbackRegex = /let orgData = \\\\{{.*?\\\\}}/;
                    updatedHTML = updatedHTML.replace(fallbackRegex, "let orgData = " + JSON.stringify(orgData) + ";");
                }}

                // Replace callListData line
                const callListDataRegex = /let callListData = \\\\[[^]*?\\\\];/;
                const newCallListDataLine = "let callListData = " + JSON.stringify(callListData) + ";";
                if (callListDataRegex.test(updatedHTML)) {{
                    updatedHTML = updatedHTML.replace(callListDataRegex, newCallListDataLine);
                }} else {{
                    const fallbackRegex = /let callListData = \\\\[.*?\\\\];/;
                    updatedHTML = updatedHTML.replace(fallbackRegex, "let callListData = " + JSON.stringify(callListData) + ";");
                }}

                const encodedContent = btoa(unescape(encodeURIComponent(updatedHTML)));

                const putRes = await fetch(`https://api.github.com/repos/${{owner}}/${{repo}}/contents/${{path}}`, {{
                    method: "PUT",
                    headers: {{
                        "Authorization": `token ${{pat}}`,
                        "Accept": "application/vnd.github.v3+json",
                        "Content-Type": "application/json"
                    }},
                    body: JSON.stringify({{
                        message: "Auto-sync contact list and org chart edits (via Browser GUI)",
                        content: encodedContent,
                        sha: sha,
                        branch: branch
                    }})
                }});

                if (!putRes.ok) {{
                    throw new Error(`Commit failed (${{putRes.status}})`);
                }}

                showToast("Saved & Synced to GitHub!", "success");
                localStorage.removeItem(LS_KEY);

            }} catch (error) {{
                console.error("Auto-sync failed:", error);
                showToast("Auto-sync failed: " + error.message, "error");
            }}
        }}

        let lastKnownFileSha = null;
        let updateCheckInterval = null;

        async function checkGitHubForUpdates() {{
            // Safety guard: if user is currently editing a contact, don't auto-update to avoid losing their typed input
            const editMode = document.getElementById("edit-mode-container");
            if (editMode && editMode.style.display === "block") {{
                return;
            }}

            const owner = localStorage.getItem(GH_KEY_PREFIX + "owner") || "jreezycalhigh";
            const repo = localStorage.getItem(GH_KEY_PREFIX + "repo") || "SAP-Org-Chart";
            const path = localStorage.getItem(GH_KEY_PREFIX + "path") || "SAP_Account_Priority_Outreach_Hub.html";
            const branch = localStorage.getItem(GH_KEY_PREFIX + "branch") || "main";
            const pat = localStorage.getItem(GH_KEY_PREFIX + "pat") || "";

            const headers = {{
                "Accept": "application/vnd.github.v3+json",
                "Cache-Control": "no-cache"
            }};
            if (pat) {{
                headers["Authorization"] = `token ${{pat}}`;
            }}

            try {{
                const url = `https://api.github.com/repos/${{owner}}/${{repo}}/contents/${{path}}?ref=${{branch}}`;
                const res = await fetch(url, {{ headers }});
                if (!res.ok) return;

                const data = await res.json();
                const currentSha = data.sha;

                if (!lastKnownFileSha) {{
                    lastKnownFileSha = currentSha;
                    return;
                }}

                if (currentSha !== lastKnownFileSha) {{
                    console.log("Detecting new remote changes on GitHub. Syncing real-time...");
                    
                    const b64Raw = data.content.replace(/\s/g, "");
                    const remoteHTML = decodeURIComponent(escape(atob(b64Raw)));

                    const orgDataMatch = remoteHTML.match(/let orgData = (\\{{[^]*?\\}});/);
                    const callListDataMatch = remoteHTML.match(/let callListData = (\\\\[[^]*?\\\\]);/);

                    if (orgDataMatch && callListDataMatch) {{
                        const newOrgData = JSON.parse(orgDataMatch[1]);
                        const newCallListData = JSON.parse(callListDataMatch[1]);

                        orgData = newOrgData;
                        callListData = newCallListData;

                        renderOrgTree();
                        renderCallList();
                        renderScheduleTab();

                        showToast("Updated with live remote changes!", "success");
                        lastKnownFileSha = currentSha;
                    }}
                }}
            }} catch (err) {{
                console.warn("Real-time update check failed:", err);
            }}
        }}

        function startRealTimeSync() {{
            if (updateCheckInterval) clearInterval(updateCheckInterval);
            setTimeout(checkGitHubForUpdates, 5000);
            updateCheckInterval = setInterval(checkGitHubForUpdates, 30000);
        }}

        // Auto-run loadSyncCredentials on startup
        window.addEventListener("DOMContentLoaded", () => {{
            updateGitHubSyncButtonState();
            startRealTimeSync();
        }});

    </script>

</body>

</html>

"""

with open("SAP_Account_Priority_Outreach_Hub.html", "w", encoding="utf-8") as f:

    f.write(html_content)

print("Generated SAP_Account_Priority_Outreach_Hub.html successfully!")

